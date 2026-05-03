import streamlit as st
import pandas as pd
import json
import io
import os
import re
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from difflib import SequenceMatcher
from google import genai
from google.genai import types as genai_types
from openpyxl import load_workbook

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="مهووس | معالج المنتجات الذكي",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    body, .stApp { direction: rtl; font-family: 'Segoe UI', Tahoma, Arial, sans-serif; }
    .brand-card {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border: 1px solid #0f3460;
        border-radius: 12px;
        padding: 20px;
        color: white;
        margin-bottom: 16px;
    }
    .stat-box {
        background: rgba(255,255,255,0.05);
        border-radius: 8px;
        padding: 12px;
        text-align: center;
    }
    .section-header {
        background: linear-gradient(90deg, #0f3460, #533483);
        color: white;
        padding: 10px 16px;
        border-radius: 8px;
        font-weight: bold;
        margin: 12px 0;
    }
</style>
""", unsafe_allow_html=True)

# ─── COMPETITOR STORES ───────────────────────────────────────────────────────
COMPETITOR_STORES = [
    "https://saeedsalah.com/", "https://vanilla.sa/", "https://sara-makeup.com/",
    "https://alkhabeershop.com/", "https://www.goldenscent.com/", "https://leesanto.com/",
    "https://azalperfume.com/", "https://candyniche.com/", "https://luxuryperfumesnish.com/",
    "https://hanan-store55.com/", "https://areejamwaj.com/", "https://niceonesa.com/",
    "https://www.sephora.me/sa-ar", "https://www.faces.sa/ar", "https://niche.sa/",
    "https://worldgivenchy.com/", "https://sarahmakeup37.com/", "https://aromaticcloud.com/",
    "https://tatayab.com/", "https://kayan9.com/",
    "https://www.noon.com/saudi-ar/", "https://www.amazon.sa/",
    "https://en.ounass.com/saudi-arabia/", "https://www.namshi.com/sa-ar/",
    "https://www.brandsforless.com/en-sa/", "https://www.sivvi.com/en-sa/",
    "https://haraj.com.sa/", "https://shukran.com/",
]

# ─── HTML TEMPLATES ──────────────────────────────────────────────────────────
HTML_TEMPLATE_NEW = (
    '<p><strong>[مقدمة تسويقية جذابة]</strong></p>'
    '<h3>التفاصيل والخطوط العطرية باختصار</h3>'
    '<ul>'
      '<li><strong>الماركة:</strong> [اسم الماركة]</li>'
      '<li><strong>اسم العطر:</strong> [اسم العطر]</li>'
      '<li><strong>الجنس:</strong> [رجالي / نسائي / للجنسين]</li>'
      '<li><strong>الخط العطري (العائلة):</strong> [مثل: حمضي - أروماتك]</li>'
      '<li><strong>الحجم:</strong> [الحجم مل]</li>'
      '<li><strong>التركيز:</strong> [مثل: أودي بارفيوم]</li>'
      '<li><strong>سنة الإصدار:</strong> [السنة والتوقيع إن وجد]</li>'
    '</ul>'
    '<h3>رحلة العطر (النوتات)</h3>'
    '<ul>'
      '<li><strong>الافتتاحية:</strong> [وصف]</li>'
      '<li><strong>القلب العطري:</strong> [وصف]</li>'
      '<li><strong>القاعدة الأساسية:</strong> [وصف]</li>'
    '</ul>'
    '<h3>لماذا تختار هذا العطر؟</h3>'
    '<ul>'
      '<li><strong>رائحة متوازنة:</strong> [وصف]</li>'
      '<li><strong>مثالي لجميع الأوقات:</strong> [وصف]</li>'
      '<li><strong>أداء قوي:</strong> [وصف]</li>'
    '</ul>'
    '<h3>الأسئلة الشائعة</h3>'
    '<p><strong>هل هذا العطر مناسب للمناخ الحار؟</strong><br>[إجابة]</p>'
    '<p><strong>هل يمكن استخدامه يومياً؟</strong><br>[إجابة]</p>'
    '<p><strong>ما هي المناسبة الأفضل لاستخدامه؟</strong><br>[إجابة]</p>'
    '<h3>اكتشف المزيد من مهووس</h3>'
    '<ul>'
      '<li><a href="/categories/perfumes-men">استكشف أحدث العطور الرجالية</a></li>'
      '<li><a href="/categories/perfumes-women">تصفح أجمل العطور النسائية الجذابة</a></li>'
      '<li><a href="/categories/niche-perfumes">للباحثين عن التميز، استكشف عطور النيش الفاخرة</a></li>'
    '</ul>'
)

HTML_TEMPLATE_TESTER = (
    '<p><strong>استمتع بالفخامة المطلقة بتكلفة أذكى! نقدم لك تستر "[اسم العطر]" '
    'من [الماركة] الأصلي 100%، ليمنحك نفس التجربة، الثبات، والفوحان للإصدار '
    'المغلف ولكن بسعر استثنائي. [مقدمة عن العطر].</strong></p>'
    '<h3>التفاصيل والخطوط باختصار</h3>'
    '<ul>'
      '<li><strong>الماركة:</strong> [اسم الماركة]</li>'
      '<li><strong>الاسم:</strong> [اسم العطر]</li>'
      '<li><strong>حالة المنتج:</strong> تستر (Tester) أصلي 100%.</li>'
      '<li><strong>الجنس:</strong> [رجالي / نسائي / للجنسين]</li>'
      '<li><strong>الخط (العائلة):</strong> [مثل: حمضي - أروماتك]</li>'
      '<li><strong>الحجم:</strong> [الحجم مل]</li>'
      '<li><strong>التركيز:</strong> [مثل: أودي بارفيوم]</li>'
    '</ul>'
    '<h3>رحلة النوتات</h3>'
    '<ul>'
      '<li><strong>الافتتاحية:</strong> [وصف]</li>'
      '<li><strong>القلب:</strong> [وصف]</li>'
      '<li><strong>القاعدة الأساسية:</strong> [وصف]</li>'
    '</ul>'
    '<h3>لماذا تختار هذا الإصدار؟</h3>'
    '<ul>'
      '<li><strong>رائحة متوازنة:</strong> [وصف]</li>'
      '<li><strong>مثالي لجميع الأوقات:</strong> [وصف]</li>'
      '<li><strong>أداء قوي:</strong> [وصف]</li>'
    '</ul>'
    '<h3>الدليل الشامل للتساتر من متجر مهووس</h3>'
    '<p>هل تتساءل عن سر التساتر ولماذا تحظى بشعبية هائلة بين عشاق الروائح '
    'الفاخرة؟ يسعدنا في متجر مهووس أن نكشف لك هذا السر، لنجعل تجربة تسوقك أكثر '
    'ذكاءً وثقة.</p>'
    '<p><strong>ما هو التستر؟</strong><br>التستر هو نسخة أصلية 100% تصدرها '
    'الشركة المصنعة (الماركات العالمية) جنباً إلى جنب مع المنتجات التجارية. '
    'الهدف الأساسي من إنتاجه هو وضعه في المتاجر والبوتيكات الفاخرة ليتمكن '
    'العملاء من تجربة الرائحة والأداء قبل الشراء.</p>'
    '<p><strong>ما الفرق بين التستر والإصدار العادي المغلف؟</strong><br>'
    'الفرق الوحيد والأساسي يكمن في "الشكل الخارجي" فقط، ولا مساومة أبداً على '
    'الجودة:</p>'
    '<ul>'
      '<li><strong>السائل:</strong> متطابق 100% من حيث المكونات، التركيز، '
      'الثبات، والفوحان. أنت تحصل على نفس القطرة الأصلية تماماً.</li>'
      '<li><strong>الزجاجة:</strong> يأتي في نفس الزجاجة الأصلية الفاخرة '
      'للماركة، وقد يُطبع عليها أحياناً عبارة (Tester) أو (Demonstration).</li>'
      '<li><strong>العلبة الخارجية:</strong> بهدف تقليل التكاليف، تُصدر '
      'الشركات التساتر في علب كرتونية بسيطة (غالباً بيضاء أو بنية صديقة '
      'للبيئة)، وتأتي بدون الغلاف البلاستيكي الشفاف (السلوفان).</li>'
      '<li><strong>الغطاء:</strong> تأتي معظم التساتر بغطائها الأصلي الفاخر، '
      'وفي حالات نادرة جداً قد تأتي بدون غطاء بناءً على تصميم الشركة '
      'المصنعة.</li>'
    '</ul>'
    '<p><strong>لماذا يعتبر التستر استثماراً ذكياً؟</strong><br>إذا كنت '
    'تشتري لاقتنائك الشخصي وليس لتقديمه كهدية رسمية، فإن التستر هو الخيار '
    'الأكثر ذكاءً وتوفيراً. فهو يتيح لك الاستمتاع بأرقى الروائح العالمية '
    'وإصدارات النيش بأسعار اقتصادية مخفضة جداً، لتحصل على أقصى قيمة مقابل ما '
    'تدفعه.</p>'
    '<p><strong>ضمان مهووس الذهبي</strong><br>نحن في متجر مهووس نضع ثقتك في '
    'المقام الأول. نضمن لك أصالة جميع التساتر المتوفرة لدينا بنسبة 100%. يتم '
    'توفيرها من نفس الموزعين المعتمدين للماركات العالمية، لتعيش تجربة الفخامة '
    'المطلقة براحة بال تامة.</p>'
    '<h3>اكتشف المزيد من مهووس</h3>'
    '<ul>'
      '<li><a href="/categories/testers">تصفح تشكيلتنا الواسعة من التساتر '
      'الأصلية</a></li>'
      '<li><a href="/categories/niche-perfumes-men">تسوق المزيد من إصدارات '
      'النيش الرجالية الفاخرة</a></li>'
      '<li><a href="/categories/niche-perfumes-women">اكتشف أحدث إصدارات '
      'النيش النسائية</a></li>'
    '</ul>'
)


def minify_html(html: str) -> str:
    """يُزيل الفراغات الزائدة بين عناصر HTML لمنع ظهور أسطر فارغة في عرض سلة.

    يحوّل: '<p>...</p>\n  <h3>...' إلى: '<p>...</p><h3>...'
    """
    if not html:
        return ''
    # احذف الفراغات بين الوسوم
    cleaned = re.sub(r'>\s+<', '><', str(html).strip())
    # احذف الأسطر الجديدة وعلامات Tab والمسافات المتعددة داخل النص
    cleaned = re.sub(r'[\r\n\t]+', ' ', cleaned)
    cleaned = re.sub(r'  +', ' ', cleaned)
    return cleaned.strip()


# ═════════════════════════════════════════════════════════════════════════════
#  تصنيفات التستر — مهووس
# ═════════════════════════════════════════════════════════════════════════════

TESTER_CATEGORY_MEN    = 'العطور > عطور التستر > عطور التستر رجالية'
TESTER_CATEGORY_WOMEN  = 'العطور > عطور التستر > عطور التستر نسائية'
TESTER_CATEGORY_NICHE  = 'العطور > عطور التستر > عطور التستر النيش'
TESTER_CATEGORY_NEW    = 'العطور > عطور التستر > عطور التستر جديدة'
TESTER_CATEGORY_ROOT   = 'العطور > عطور التستر'


def map_to_tester_category(base_category: str, perfume_name: str = '') -> str:
    """يُحوّل تصنيف العطر الأساسي إلى تصنيف التستر المقابل.

    أمثلة:
      'العطور > عطور رجالية > رسمية'  → 'العطور > عطور التستر > عطور التستر رجالية'
      'العطور > عطور النيش > للجنسين' → 'العطور > عطور التستر > عطور التستر النيش'
      'العطور > عطور نسائية > جذابة'  → 'العطور > عطور التستر > عطور التستر نسائية'
    """
    bc = str(base_category or '')
    pn = str(perfume_name or '').lower()
    pn_norm = re.sub(r'[أإآ]', 'ا', pn)

    # التحقق أولاً من النيش (له أولوية عالية لأن الإصدارات النيش غالباً unisex)
    if 'النيش' in bc or 'نيش' in bc or 'niche' in pn:
        return TESTER_CATEGORY_NICHE

    # استخدم detect_gender المحسّنة (تحتوي قائمة كلمات أوسع)
    gender = detect_gender(perfume_name, bc)
    if gender == 'رجالي':
        return TESTER_CATEGORY_MEN
    if gender == 'نسائي':
        return TESTER_CATEGORY_WOMEN
    # للجنسين → نيش (لأن detect_gender يُرجع للجنسين فقط للنيش)
    return TESTER_CATEGORY_NICHE


def detect_gender(perfume_name: str, category: str = '') -> str:
    """يستنتج جنس العطر من الاسم/التصنيف."""
    pn = str(perfume_name or '').lower()
    pn_norm = re.sub(r'[أإآ]', 'ا', pn)
    bc = str(category or '')

    # 1) نيش للجنسين / unisex صريح
    if 'unisex' in pn or 'للجنسين' in pn_norm or 'للجنسين' in bc:
        return 'للجنسين'

    # 2) نسائي — كلمات صريحة
    feminine_keywords = [
        'نسائ', 'للنساء', 'للسيدات', 'سيدات',
        'women', 'femme', 'pour femme', 'lady', 'her', 'miss',
        'mademoiselle', 'queen', 'princess',
        'ليدي',  # Paco Rabanne Lady Million
        'كوين',  # Queen
        'مدام',
    ]
    for kw in feminine_keywords:
        if kw in pn_norm or (' ' + kw + ' ') in (' ' + pn + ' '):
            return 'نسائي'

    # 3) رجالي — كلمات صريحة
    masculine_keywords = [
        'رجال', 'للرجال', 'رجالي',
        ' men ', ' men', 'homme', 'pour homme', 'mr ', 'mr.',
        'gentleman', 'boss', 'king',
    ]
    for kw in masculine_keywords:
        if kw in pn_norm or (' ' + kw + ' ') in (' ' + pn + ' '):
            return 'رجالي'

    # 4) من التصنيف الأساسي
    if 'نسائ' in bc:
        return 'نسائي'
    if 'رجال' in bc:
        return 'رجالي'

    # 5) إذا لم نتمكن من تحديد، نختار "رجالي" كافتراضي للعطور التجارية
    # (الجزء الأكبر من السوق رجالي إلا إذا كان نيش)
    if 'النيش' in bc or 'niche' in pn:
        return 'للجنسين'
    return 'رجالي'


def detect_concentration(perfume_name: str) -> str:
    """يستنتج تركيز العطر من اسمه (Eau de Parfum / Toilette / Cologne / Elixir / Parfum)."""
    pn = str(perfume_name or '').lower()
    pn_norm = re.sub(r'[أإآ]', 'ا', pn)

    # Elixir / إكسير / اليكسير (الأكثر تركيزاً)
    if 'elixir' in pn or 'اليكسير' in pn_norm or 'الكسير' in pn_norm or 'اكسير' in pn_norm:
        return 'إكسير (Elixir)'
    # Parfum / Extrait
    if ('extrait' in pn or 'parfum intense' in pn
            or 'برفان انتنس' in pn_norm or 'انتنس' in pn_norm and 'parfum' in pn):
        return 'إكستريت دي بارفان (Extrait)'
    # Eau de Parfum / EDP
    if (' edp' in f' {pn} ' or 'eau de parfum' in pn or 'بارفيوم' in pn_norm
            or 'بارفان' in pn_norm or 'برفان' in pn_norm or 'برفيوم' in pn_norm
            or 'parfum' in pn):
        return 'أو دو بارفيوم (EDP)'
    # Eau de Toilette / EDT
    if (' edt' in f' {pn} ' or 'eau de toilette' in pn or 'تواليت' in pn_norm
            or 'توالت' in pn_norm or 'toilette' in pn):
        return 'أو دو تواليت (EDT)'
    # Eau de Cologne
    if 'cologne' in pn or 'كولونيا' in pn_norm:
        return 'أو دو كولون (EDC)'

    return 'أو دو بارفيوم'


def clean_perfume_display_name(name: str) -> str:
    """يُزيل كلمات «عطر/تستر» من الاسم لعرضه في الوصف بشكل احترافي."""
    if not name:
        return ''
    n = str(name).strip()
    # احذف «تستر» من النهاية أو البداية
    n = re.sub(r'^(?:تستر|تيستر|تستير|tester|testr)\s+', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s+(?:تستر|تيستر|تستير|tester|testr)\s*$', '', n, flags=re.IGNORECASE)
    # احذف «عطر/العطر» في أي مكان (بداية، وسط، نهاية)
    n = re.sub(r'\bعطر\b', '', n)
    n = re.sub(r'\bالعطر\b', '', n)
    # نظّف الفراغات الزائدة
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def clean_brand_display_name(brand: str) -> str:
    """يُنظّف اسم الماركة من الفواصل/الترميز.

    أمثلة:
      'باكو رابان | Paco Rabanne' → 'باكو رابان'  (الاسم العربي فقط للعرض)
      'Paco Rabanne | باكو رابان' → 'باكو رابان'
      'كرييد' → 'كرييد'
    """
    if not brand:
        return ''
    b = str(brand).strip()
    # إذا كانت الصيغة "عربي | إنجليزي" أو العكس، اختر العربية للعرض في النص
    if '|' in b:
        parts = [p.strip() for p in b.split('|') if p.strip()]
        # ابحث عن الجزء الذي يحتوي حروفاً عربية
        ar_part = next(
            (p for p in parts if re.search(r'[\u0600-\u06FF]', p)),
            None,
        )
        if ar_part:
            return ar_part
        # وإلا، خذ أول جزء
        return parts[0] if parts else b
    return b


def fill_tester_template_basics(html_template: str, brand_name: str,
                                 perfume_name: str, size_ml: int,
                                 base_category: str = '') -> str:
    """يملأ ما يمكن ملؤه برمجياً في القالب: الماركة، الاسم، الحجم، الجنس، التركيز.

    الحقول التي لا يمكن ملؤها بدقة (النوتات، العائلة العطرية، الأسئلة الشائعة)
    تُترك كما هي لتُملأ لاحقاً عبر Gemini أو تُستبدل برسالة عامة.
    """
    if not html_template:
        return ''

    h = html_template
    clean_name = clean_perfume_display_name(perfume_name)
    clean_brand = clean_brand_display_name(brand_name)
    gender = detect_gender(perfume_name, base_category)
    concentration = detect_concentration(perfume_name)

    # استبدال القيم الأساسية المعروفة
    h = h.replace('[اسم الماركة]', clean_brand)
    h = h.replace('[اسم العطر]', clean_name)
    h = h.replace('[الماركة]', clean_brand)
    h = h.replace('[الحجم مل]', f'{size_ml} مل' if size_ml else '')
    h = h.replace('[رجالي / نسائي / للجنسين]', gender)
    h = h.replace('[رجالي / نسائي]', gender)
    h = h.replace('[مثل: أودي بارفيوم]', concentration)
    h = h.replace('[مثل: أو دو بارفيوم]', concentration)
    h = h.replace('[التركيز]', concentration)

    return h


# ═════════════════════════════════════════════════════════════════════════════
#  إثراء التساتر التلقائية — استدعاء Gemini لتعبئة التفاصيل المتخصصة
# ═════════════════════════════════════════════════════════════════════════════

ENRICHMENT_PROMPT = """أنت خبير عطور محترف. أمامك قائمة من تساتر العطور تحتاج لتعبئة تفاصيل احترافية لكل واحد.

المعلومات المتوفرة لكل تستر:
- اسم العطر الأساسي
- الماركة
- الحجم
- الجنس (مستنتج)
- التركيز (مستنتج)

مطلوب منك لكل تستر:
١) **مقدمة تسويقية جذابة** (سطر إلى سطرين)
٢) **العائلة العطرية** (مثل: حمضي - أروماتك / شرقي - زهري)
٣) **النوتات الثلاث**: الافتتاحية، القلب، القاعدة (وصف مختصر لكل واحد)
٤) **3 أسباب لاختيار الإصدار** (جملة واحدة لكل سبب)
٥) **3 إجابات للأسئلة الشائعة**: المناخ الحار، الاستخدام اليومي، المناسبة الأفضل

أرجع النتائج كـ JSON صحيح بالتنسيق التالي بالضبط:

```json
{{
  "enrichments": [
    {{
      "tester_id": "<معرّف التستر الذي أُعطي لك>",
      "intro": "مقدمة تسويقية...",
      "fragrance_family": "شرقي - خشبي",
      "top_notes": "البرغموت والمندرين...",
      "heart_notes": "الياسمين والورد...",
      "base_notes": "العنبر والمسك...",
      "reason_1": "رائحة فاخرة طويلة الأمد",
      "reason_2": "مناسب للمناسبات الرسمية والكاجوال",
      "reason_3": "ثبات قوي يدوم طوال اليوم",
      "faq_hot": "نعم، يتميز بأداء قوي في المناخ الحار...",
      "faq_daily": "مثالي للاستخدام اليومي بفضل توازنه...",
      "faq_occasion": "يناسب المناسبات المسائية والرسمية..."
    }}
  ]
}}
```

🔴 قواعد صارمة:
- استخدم **معلومات حقيقية وصحيحة** عن كل عطر — لا تخترع نوتات
- إذا لم تكن متأكداً من النوتات الفعلية لعطر معين، اكتب وصفاً عاماً موثوقاً مرتبطاً بنوع العطر وعائلته
- اكتب باللغة العربية فقط
- لا تُضِف أي حقول إضافية خارج التنسيق المطلوب
- المعرّف tester_id يجب أن يطابق المعطى لك بالحرف

قائمة التساتر:
{testers_json}
"""


def enrich_auto_added_testers(
    result: dict,
    brand_name: str,
    api_key: str,
    model_name: str = 'gemini-2.5-flash',
    batch_size: int = 8,
    progress_cb=None,
) -> dict:
    """يستدعي Gemini لتعبئة تفاصيل احترافية للتساتر التلقائية.

    لكل تستر مع `_auto_added=True`:
      - يستبدل [وصف]، [الخط العطري]، [الافتتاحية]، إلخ. بمعلومات حقيقية
      - يحدّث `seo_description` بمقدمة احترافية

    Args:
        result: قاموس النتائج بعد ensure_all_testers_added
        brand_name: اسم الماركة (للسياق)
        api_key: مفتاح Gemini API
        model_name: اسم الموديل
        batch_size: عدد التساتر في كل استدعاء (افتراضي 8)
        progress_cb: callback اختياري(int, int) لتحديث التقدم

    Returns:
        result مُحدَّث بأوصاف كاملة. إذا فشل Gemini، يُترك القالب كما هو.
    """
    if not isinstance(result, dict):
        return result

    auto_testers = [
        t for t in (result.get('testers_to_add') or [])
        if isinstance(t, dict) and t.get('_auto_added')
    ]
    if not auto_testers:
        return result

    try:
        client = genai.Client(api_key=api_key)
    except Exception:
        return result  # بدون Gemini، نُرجع كما هو

    config = genai_types.GenerateContentConfig(
        temperature=0.4,
        max_output_tokens=16384,
        response_mime_type='application/json',
    )

    total = len(auto_testers)
    processed = 0

    for batch_start in range(0, total, batch_size):
        batch = auto_testers[batch_start:batch_start + batch_size]

        # بناء قائمة JSON للدفعة
        batch_data = []
        for idx, t in enumerate(batch):
            tester_id = f"t{batch_start + idx}"
            t['_enrich_id'] = tester_id  # نستخدم هذا للتطابق
            clean_name = clean_perfume_display_name(t.get('name', ''))
            batch_data.append({
                'tester_id': tester_id,
                'perfume_name': clean_name,
                'brand': brand_name,
                'size_ml': t.get('size_ml', 100),
                'gender': detect_gender(t.get('name', '')),
                'concentration': detect_concentration(t.get('name', '')),
            })

        prompt = ENRICHMENT_PROMPT.format(
            testers_json=json.dumps(batch_data, ensure_ascii=False, indent=2)
        )

        # استدعاء Gemini مع إعادة محاولات بسيطة
        text = ''
        for attempt in range(3):
            try:
                resp = client.models.generate_content(
                    model=model_name,
                    contents=prompt,
                    config=config,
                )
                text = (resp.text or '').strip()
                if text:
                    break
            except Exception:
                if attempt < 2:
                    time.sleep(2 * (attempt + 1))
                continue

        if not text:
            # فشل الإثراء لهذه الدفعة — تخطّ وكمل
            processed += len(batch)
            if progress_cb:
                try:
                    progress_cb(processed, total)
                except Exception:
                    pass
            continue

        # تحليل JSON
        try:
            data = extract_json(text)
        except Exception:
            processed += len(batch)
            if progress_cb:
                try:
                    progress_cb(processed, total)
                except Exception:
                    pass
            continue

        enrichments = data.get('enrichments') or []
        if not isinstance(enrichments, list):
            enrichments = []

        # طبّق الإثراء على كل تستر
        enrich_by_id = {
            str(e.get('tester_id', '') or ''): e
            for e in enrichments if isinstance(e, dict)
        }

        for t in batch:
            tester_id = t.get('_enrich_id', '')
            e = enrich_by_id.get(tester_id)
            if not e:
                continue

            # ابنِ الوصف من القالب الكامل مع كل القيم المُملوءة
            new_desc = fill_tester_template_complete(
                brand_name=brand_name,
                perfume_name=t.get('name', ''),
                size_ml=t.get('size_ml', 100),
                base_category='',  # سيُحدَّد في build_output_excel
                enrichment=e,
            )
            t['new_description'] = new_desc

            # حدّث seo_description بالمقدمة الجديدة
            intro = (e.get('intro') or '').strip()
            if intro:
                t['seo_description'] = intro[:155]

            # احذف معرّف الإثراء بعد الانتهاء
            t.pop('_enrich_id', None)

        processed += len(batch)
        if progress_cb:
            try:
                progress_cb(processed, total)
            except Exception:
                pass

    return result


def fill_tester_template_complete(
    brand_name: str,
    perfume_name: str,
    size_ml: int,
    base_category: str = '',
    enrichment: dict = None,
) -> str:
    """يبني وصف تستر كامل بكل التفاصيل المملوءة، جاهز للعرض على العميل."""
    enrichment = enrichment or {}
    clean_name = clean_perfume_display_name(perfume_name)
    clean_brand = clean_brand_display_name(brand_name)
    gender = detect_gender(perfume_name, base_category)
    concentration = detect_concentration(perfume_name)

    intro = (enrichment.get('intro') or
             f"تستر {clean_name} من {clean_brand} الأصلي 100% — نفس السائل والثبات والفوحان للإصدار المغلف بسعر استثنائي.")
    family = enrichment.get('fragrance_family') or 'شرقي - خشبي'
    top = enrichment.get('top_notes') or 'افتتاحية متوازنة وجذابة'
    heart = enrichment.get('heart_notes') or 'قلب عطري دافئ ومتناغم'
    base = enrichment.get('base_notes') or 'قاعدة ثابتة طويلة الأمد'
    r1 = enrichment.get('reason_1') or 'رائحة فاخرة بصبغة مميزة'
    r2 = enrichment.get('reason_2') or 'ثبات قوي وفوحان ممتاز يدوم طويلاً'
    r3 = enrichment.get('reason_3') or 'مناسب لمختلف المناسبات والأوقات'
    faq_hot = (enrichment.get('faq_hot') or
               'نعم، يتميز هذا الإصدار بأداء جيد في المناخ الحار بفضل تركيبته المتوازنة.')
    faq_daily = (enrichment.get('faq_daily') or
                 'مثالي للاستخدام اليومي وللمناسبات الخاصة على حد سواء.')
    faq_occasion = (enrichment.get('faq_occasion') or
                    'يناسب المناسبات الرسمية والكاجوال والخروج المسائي.')

    html = (
        f'<p><strong>{intro}</strong></p>'
        f'<h3>التفاصيل والخطوط باختصار</h3>'
        f'<ul>'
          f'<li><strong>الماركة:</strong> {clean_brand}</li>'
          f'<li><strong>الاسم:</strong> {clean_name}</li>'
          f'<li><strong>حالة المنتج:</strong> تستر (Tester) أصلي 100%.</li>'
          f'<li><strong>الجنس:</strong> {gender}</li>'
          f'<li><strong>الخط (العائلة):</strong> {family}</li>'
          f'<li><strong>الحجم:</strong> {size_ml} مل</li>'
          f'<li><strong>التركيز:</strong> {concentration}</li>'
        f'</ul>'
        f'<h3>رحلة النوتات</h3>'
        f'<ul>'
          f'<li><strong>الافتتاحية:</strong> {top}</li>'
          f'<li><strong>القلب:</strong> {heart}</li>'
          f'<li><strong>القاعدة الأساسية:</strong> {base}</li>'
        f'</ul>'
        f'<h3>لماذا تختار هذا الإصدار؟</h3>'
        f'<ul>'
          f'<li><strong>رائحة متوازنة:</strong> {r1}</li>'
          f'<li><strong>مثالي لجميع الأوقات:</strong> {r2}</li>'
          f'<li><strong>أداء قوي:</strong> {r3}</li>'
        f'</ul>'
        f'<h3>الأسئلة الشائعة</h3>'
        f'<p><strong>هل هذا التستر مناسب للمناخ الحار؟</strong><br>{faq_hot}</p>'
        f'<p><strong>هل يمكن استخدامه يومياً؟</strong><br>{faq_daily}</p>'
        f'<p><strong>ما هي المناسبة الأفضل لاستخدامه؟</strong><br>{faq_occasion}</p>'
        f'<h3>الدليل الشامل للتساتر من متجر مهووس</h3>'
        f'<p>هل تتساءل عن سر التساتر ولماذا تحظى بشعبية هائلة بين عشاق الروائح '
        f'الفاخرة؟ يسعدنا في متجر مهووس أن نكشف لك هذا السر، لنجعل تجربة تسوقك '
        f'أكثر ذكاءً وثقة.</p>'
        f'<p><strong>ما هو التستر؟</strong><br>التستر هو نسخة أصلية 100% '
        f'تصدرها الشركة المصنعة (الماركات العالمية) جنباً إلى جنب مع المنتجات '
        f'التجارية. الهدف الأساسي من إنتاجه هو وضعه في المتاجر والبوتيكات '
        f'الفاخرة ليتمكن العملاء من تجربة الرائحة والأداء قبل الشراء.</p>'
        f'<p><strong>ما الفرق بين التستر والإصدار العادي المغلف؟</strong><br>'
        f'الفرق الوحيد يكمن في "الشكل الخارجي" فقط، ولا مساومة على الجودة:</p>'
        f'<ul>'
          f'<li><strong>السائل:</strong> متطابق 100% من حيث المكونات، التركيز، '
          f'الثبات، والفوحان.</li>'
          f'<li><strong>الزجاجة:</strong> نفس الزجاجة الأصلية الفاخرة، وقد '
          f'يُطبع عليها أحياناً عبارة (Tester).</li>'
          f'<li><strong>العلبة الخارجية:</strong> علب كرتونية بسيطة لتقليل '
          f'التكاليف.</li>'
          f'<li><strong>الغطاء:</strong> غطاء أصلي فاخر في الغالب.</li>'
        f'</ul>'
        f'<p><strong>ضمان مهووس الذهبي</strong><br>نضمن أصالة جميع التساتر '
        f'بنسبة 100% — من نفس الموزعين المعتمدين للماركات العالمية.</p>'
        f'<h3>اكتشف المزيد من مهووس</h3>'
        f'<ul>'
          f'<li><a href="/categories/testers">تصفح تشكيلتنا الواسعة من '
          f'التساتر الأصلية</a></li>'
          f'<li><a href="/categories/niche-perfumes-men">تسوق المزيد من '
          f'إصدارات النيش الرجالية الفاخرة</a></li>'
          f'<li><a href="/categories/niche-perfumes-women">اكتشف أحدث '
          f'إصدارات النيش النسائية</a></li>'
        f'</ul>'
    )
    return html

SYSTEM_INSTRUCTION_TEMPLATE = """## هويتك ومهمتك
أنت **خبير عطور محترف بخبرة 20 سنة** + محلل تنافسي لمتجر مهووس في السوق السعودي.
مهمتك:
١) **إضافة تستر إلزامي** لكل عطر أساسي ليس لديه تستر في قائمتنا (سياسة جديدة).
٢) اكتشاف المنتجات الناقصة عند المنافسين.
٣) كتابة وصف احترافي **فقط للمنتجات الجديدة** التي تقترحها.

## ❌ ممنوع منعاً باتاً
- لا تُعيد كتابة أو تُحدّث وصف أي منتج موجود مسبقاً في قائمتنا
- products_updated يجب أن يكون دائماً [] قائمة فارغة
- لا تقترح منتجاً موجوداً مسبقاً ولو بصيغة مختلفة

## قواعد التساتر (الأهم) — السياسة الجديدة: التستر إلزامي لكل عطر *مؤهل*

**🛑 استثناءات مهمة (لا تُضِف تستراً لها أبداً):**
1. **الأطقم/المجموعات**: أي منتج يحتوي اسمه على «طقم/أطقم/مجموعة/Set/Kit/Bundle/Collection/Box» أو «N قطع» (مثل: "3 قطع") أو رمز «+» الذي يدل على دمج عدة منتجات. مثال يجب تجاهله: "مجموعة باكو رابان مليون جولد للنساء 3 قطع".
2. **الأحجام الصغيرة المكررة**: إذا كان نفس العطر موجوداً بأحجام متعددة (مثل 100مل و50مل لنفس الإصدار)، أضف تستراً **فقط للحجم الأكبر**. تجاهل الأحجام الأصغر تماماً.

**القاعدة 1 — فحص وجود التستر:**
- لكل عطر أساسي *مؤهل* في قائمتنا (ليس طقماً، وهو الحجم الأكبر لإصداره)، تحقق: هل يوجد منتج آخر في القائمة يحتوي اسمه على "تستر" أو "Tester" لنفس العطر؟
- إذا وُجد التستر → **تخطّ، لا تقترح شيئاً**
- إذا لم يُوجد → انتقل للقاعدة 2 (إلزامي).

**القاعدة 2 — البحث المرجعي عن سعر التستر + الإضافة الإلزامية:**
- ابحث في Google عن سعر التستر عند المنافسين بهذه الصيغ:
  1. "[اسم العطر] tester Saudi Arabia price"
  2. "[اسم العطر] تستر السعر السعودية"
  3. "[اسم الماركة] tester site:sa"
- ابحث في **أي** متجر سعودي يظهر في النتائج (سواء كان في القائمة المرجعية أم لا)
- سجّل سعر المنافس في حقل `competitor_price` (للمرجعية فقط) واسم المتجر في `source_store`
- 🔴 **سياسة إلزامية:** أضف التستر في `testers_to_add` **دائماً** لكل عطر أساسي مؤهل ليس له تستر في قائمتنا، **حتى لو لم تجد التستر عند أي منافس سعودي**:
  - إذا وجدت سعر المنافس → ضعه في `competitor_price`، وضع `tester_available_in_market = true`، واسم المتجر في `source_store`.
  - إذا لم تجد → اجعل `competitor_price = 0`، `tester_available_in_market = false`، و `source_store = ""`.
  - في كل الحالات، `new_price` يُحسب بقاعدة التسعير الداخلية (موضحة أدناه) وليس بسعر المنافس.

**القاعدة 3 — اسم التستر:**
- يجب أن يبدأ اسم التستر بكلمة «تستر» (وليس في النهاية).
- صيغة الاسم: `تستر [اسم العطر بدون كلمة "عطر" في أوله]`
- مثال صحيح: «تستر باكو رابان فانتوم 100مل»
- مثال خاطئ: «عطر باكو رابان فانتوم 100مل تستر»

**القاعدة 4 — صورة التستر:**
- الصورة تُؤخذ حرفياً من حقل image_url للمنتج الأساسي الموجود في قائمتنا
- لا تبحث عن صورة جديدة للتستر أبداً
- إذا كان للمنتج الأساسي أكثر من صورة (مفصولة بفاصلة)، **انسخها كلها كما هي بفواصلها** — Salla يدعم صور متعددة.

**القاعدة 5 — التساتر بلا عطر أساسي (Orphan Testers):**
- مرّ على كل تستر في قائمتنا
- تحقق: هل يوجد منتج أساسي (غير تستر) بنفس الاسم؟
- إذا لم يوجد → أضفه في missing_products مع وصفه كعطر أساسي جديد، واذكر التستر اليتيم في orphan_testers.

## قواعد المنتجات الناقصة
- قارن قائمتنا الكاملة بما يبيعه المنافسون لنفس الماركة
- ركّز على: الأكثر مبيعاً، الإصدارات الجديدة، والأحجام المختلفة الشائعة
- كل مقترح يجب أن يكون متوفراً للشراء الآن في متجر محدد (اذكر المتجر)
- الأولوية للمنتجات الأكثر مبيعاً (bestsellers)

## أسلوب الكتابة — تعلّم من هذه الأمثلة الحقيقية
{writing_dna}

## قوالب HTML الإلزامية للمنتجات الجديدة فقط
### قالب العطور الجديدة/الأساسية:
{HTML_TEMPLATE_NEW}

### قالب التساتر الجديدة:
{HTML_TEMPLATE_TESTER}

## صرامة مطلقة ضد الاختراع
- لا تخترع عطراً أو سعراً أو رابط صورة غير موجود فعلياً
- إذا لم تجد معلومة موثوقة، اترك الحقل فارغاً (لكن **لا تحذف** التستر — التستر إلزامي)
- المنتجات الناقصة (missing_products) يجب أن تكون موجودة فعلياً في متجر سعودي محدد
- التستر **يُضاف دائماً** حتى بدون توافره في السوق، فهذه سياستنا الداخلية

## 🚫 ممنوع التكرار الداخلي (قواعد صارمة)
- ممنوع منعاً باتاً تكرار نفس العطر أو التستر داخل المصفوفة. إذا وجدت العطر في أكثر من متجر منافس، اختر المتجر الأفضل أو الأرخص واذكره **مرة واحدة فقط**، وتجاهل البقية تماماً.
- لا تكرر نفس base_product_id في testers_to_add. كل base_product_id يظهر مرة واحدة فقط مهما تعددت المتاجر التي تبيع التستر.
- لا تكرر نفس المنتج في missing_products بصيغ مختلفة (مثل "Fame Parfum 80ml" و "فيم بارفان 80 مل") — كلها نفس المنتج، اذكره مرة واحدة فقط.

## 🏷️ قواعد source_store
- يُفضَّل أن يكون من قائمة المتاجر المرجعية competitors_json
- لكن إذا وجدت المنتج في متجر سعودي آخر موثوق، سجّله بنطاقه الفعلي (مثال: "noon.com")
- ممنوع اختراع متجر غير موجود — يجب أن يكون رابط المنتج قابلاً للتحقق
- إذا لم تجد التستر عند أي منافس، اجعل `source_store = ""` (لكن أضف التستر دائماً)

## 📐 صرامة مخطط JSON
- يجب أن تحتوي **جميع** عناصر missing_products على الحقلين image_url_1 و image_url_2 بشكل دائم. إذا لم تجد صورة ثانية، اجعل قيمتها سلسلة نصية فارغة "" — ولا تحذف المفتاح أبداً.
- جميع عناصر testers_to_add يجب أن تحتوي الحقلين `competitor_price` و `tester_available_in_market` (افتراضياً 0 و false إذا لم يُوجد التستر عند منافس).
- جميع المفاتيح المذكورة في مخطط الإخراج إلزامية في كل عنصر؛ القيم الفارغة تُمثَّل بـ "" أو 0 أو false وليس بحذف المفتاح.

## المتاجر السعودية للمقارنة (انسخ source_store حرفياً من هذه القائمة):
{competitors_json}

## قواعد التسعير الداخلية (تُطبَّق دائماً على new_price)
- تستر لمنتج أقل من 1000 ريال: `new_price = original_price - 70`
- تستر لمنتج 1000 ريال فأكثر: `new_price = original_price - 150`
- ⚠️ هذه القاعدة الداخلية هي المرجع للسعر النهائي — سعر المنافس مرجعي فقط ولا يُستخدم في `new_price`.

## تعليمات الإخراج
- JSON صارم فقط، يبدأ بـ {{ وينتهي بـ }}
- لا markdown، لا نص خارج JSON
- products_updated: [] دائماً (لا تحديث للموجود)
- جميع الأسعار بالريال السعودي
"""

# ─── COLUMN DETECTION ────────────────────────────────────────────────────────
ARABIC_COL_KEYS = {
    'name':          ['اسم المنتج', 'اسم'],
    'type':          ['نوع المنتج'],
    'category':      ['فئة المنتج', 'فئة'],
    'images':        ['صورة المنتج', 'صورة'],
    'option_name':   ['اسم خيار'],
    'option_value':  ['اسم الخيار'],
    'price':         ['سعر المنتج', 'سعر'],
    'quantity':      ['الكمية'],
    'description':   ['الوصف'],
    'accepts_orders':['هل يقبل'],
    'sku':           ['sku', 'رمز المنتج'],
    'barcode':       ['الباركود', 'رمز الباركود'],
    'brand':         ['الماركة'],
    'status':        ['حالة المنتج', 'حالة'],
}

FALLBACK_POSITIONS = {
    'name': 1, 'type': 2, 'category': 3, 'images': 4,
    'price': 7, 'quantity': 8, 'description': 9,
    'sku': 11, 'brand': 22, 'status': 24,
}


def _norm_ar(s: str) -> str:
    s = str(s).strip().lower()
    s = re.sub(r'[ً-ٰٟ]', '', s)
    s = s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا').replace('ٱ', 'ا')
    s = s.replace('ى', 'ي').replace('ئ', 'ي')
    s = s.replace('ة', 'ه')
    s = re.sub(r'\s+', ' ', s)
    return s


def find_col(df: pd.DataFrame, key: str) -> str | None:
    keywords = ARABIC_COL_KEYS.get(key, [])
    EXCLUDE = {
        'name':   ['خيار', 'option', '[1]', '[2]', '[3]'],
        'price':  ['خيار', 'option', 'تكلفه', 'مخفض'],
        'images': ['وصف صوره', 'وصف صورة'],
    }
    excludes = [_norm_ar(x) for x in EXCLUDE.get(key, [])]
    cols_norm = [(col, _norm_ar(col)) for col in df.columns]

    for kw in keywords:
        kn = _norm_ar(kw)
        for col, cn in cols_norm:
            if cn == kn and not any(x in cn for x in excludes):
                return col
    for kw in keywords:
        kn = _norm_ar(kw)
        for col, cn in cols_norm:
            if kn in cn and not any(x in cn for x in excludes):
                return col
    if key in FALLBACK_POSITIONS:
        idx = FALLBACK_POSITIONS[key]
        cols = list(df.columns)
        if idx < len(cols):
            return cols[idx]
    return None


def get_brand_col(df: pd.DataFrame) -> str | None:
    col = find_col(df, 'brand')
    if col:
        return col
    for c in df.columns:
        sample = df[c].dropna().astype(str).head(30)
        pipe_count = sample.str.contains(r'\|').sum()
        avg_len = sample.str.len().mean()
        if pipe_count > 5 and avg_len < 60:
            return c
    return None


def is_tester(name: str) -> bool:
    if not isinstance(name, str) or not name.strip():
        return False
    n = name.lower()
    if any(t in n for t in ['tester', 'testr']):
        return True
    n_norm = re.sub(r'[أإآ]', 'ا', n)
    return any(t in n_norm for t in ['تستر', 'تستير', 'تيستر'])


# كلمات تدل على «طقم/مجموعة» (لا تستحق إضافة تستر)
_SET_KEYWORDS_AR = (
    'طقم', 'أطقم', 'اطقم', 'مجموعة', 'مجموعه', 'باكج', 'باقة', 'هدية',
    'هديه', 'بوكس', 'كوليكشن',
)
_SET_KEYWORDS_EN = (
    'set ', ' set', 'gift set', 'bundle', 'collection', 'box set', 'kit',
    'pack', 'duo', 'trio',
)
# نمط للتعرف على أنماط مثل «3 قطع» / «4 قطع» / «قطعتين»
_PIECES_RX = re.compile(r'\b\d+\s*قطع\b|قطعتين|قطعتان|\b\d+\s*pcs?\b', re.IGNORECASE)


def is_set(name: str) -> bool:
    """يتحقق ما إذا كان المنتج طقماً/مجموعة (لا يحتاج تستراً).

    قواعد التحقق:
      • وجود كلمة دالة على الطقم (طقم/مجموعة/set/kit/bundle…)
      • نمط «N قطع» أو «N pcs»
      • وجود علامة «+» تدل على دمج عدة منتجات في عرض واحد
    """
    if not isinstance(name, str) or not name.strip():
        return False

    n = name.strip()
    n_lower = n.lower()
    n_ar = re.sub(r'[أإآ]', 'ا', n_lower)

    # 1) كلمات عربية
    for kw in _SET_KEYWORDS_AR:
        kw_norm = re.sub(r'[أإآ]', 'ا', kw)
        if kw_norm in n_ar:
            return True

    # 2) كلمات إنجليزية (مع حدود الكلمات لتجنب false positives في duo→ مثلاً)
    for kw in _SET_KEYWORDS_EN:
        if kw.strip() in n_lower:
            # تأكد إنها كلمة مستقلة وليست جزءاً من كلمة أكبر
            if re.search(rf'\b{re.escape(kw.strip())}\b', n_lower):
                return True

    # 3) نمط «N قطع» / «N pcs»
    if _PIECES_RX.search(n):
        return True

    # 4) علامة «+» في الاسم تدل على دمج عدة منتجات
    if '+' in n:
        return True

    return False


def _extract_size_for_grouping(name: str) -> int:
    """يستخرج الحجم بالـ مل بشكل قوي (أكبر رقم متبوع بـ مل/ml يُعتبر الحجم).

    مهم: نأخذ أكبر رقم لتفادي حالات مثل «1 مليون لاكي 100 مل» حيث «1»
    ليست حجماً.
    """
    if not name:
        return 0
    matches = _SIZE_RX.findall(str(name))
    if not matches:
        return 0
    sizes = []
    for m in matches:
        try:
            sizes.append(int(float(str(m).replace(',', '.'))))
        except (ValueError, TypeError):
            continue
    return max(sizes) if sizes else 0


def calc_tester_price(original_price: float) -> float:
    """قاعدة التسعير الداخلية للتسات: الأقل من 1000 ريال يخصم 70، والباقي 150."""
    try:
        p = float(original_price or 0)
    except (TypeError, ValueError):
        p = 0.0
    if p >= 1000:
        return max(p - 150, 0)
    return max(p - 70, 0)


def build_tester_name(base_name: str) -> str:
    """يُنشئ اسم تستر صحيح من اسم العطر الأساسي.

    قواعد التسمية المعتمدة:
      • إذا بدأ الاسم بـ «عطر » فإن «عطر» تُستبدل بـ «تستر»
      • وإلا تُوضع «تستر » في أول الاسم (وليس آخره)

    أمثلة:
      'عطر رابان مليون 90مل'  →  'تستر رابان مليون 90مل'
      'باكو رابان فانتوم 100مل' → 'تستر باكو رابان فانتوم 100مل'

    كما تُزيل أي ذيول تستر مكرّرة قد تكون مُلصقة من قبل.
    """
    if not base_name or not str(base_name).strip():
        return 'تستر'

    name = str(base_name).strip()

    # 1) إن كانت الكلمة «تستر/تيستر/تستير/Tester» مُلصقة في النهاية، احذفها
    name = re.sub(
        r'\s*(?:تستر|تيستر|تستير|tester|testr)\s*$',
        '',
        name,
        flags=re.IGNORECASE,
    ).strip()

    # إن لم يبقَ شيء (الاسم كان «تستر» وحدها)، أرجِعها
    if not name:
        return 'تستر'

    # 2) لو بدأ الاسم بـ «عطر »، استبدلها بـ «تستر»
    if re.match(r'^(?:عطر|العطر)\s+', name):
        return re.sub(r'^(?:عطر|العطر)\s+', 'تستر ', name, count=1)

    # 3) لو بدأ الاسم بـ «تستر » مسبقاً، رجّع الاسم كما هو
    if re.match(r'^(?:تستر|تيستر|تستير|tester|testr)\s+', name, re.IGNORECASE):
        return name

    # 4) خلاف ذلك، ضع «تستر » في الأول
    return f"تستر {name}"
    return max(p - 70, 0)


def load_products(file) -> pd.DataFrame:
    name = file.name.lower()
    if name.endswith('.csv'):
        for enc in ['utf-8-sig', 'utf-8', 'cp1256']:
            try:
                file.seek(0)
                df = pd.read_csv(file, encoding=enc)
                break
            except Exception:
                continue
    else:
        file.seek(0)
        df = pd.read_excel(file, header=1)
        first_col = str(df.columns[0])
        if first_col not in ('No.',) and not any(k in first_col for k in ['اسم', 'No']):
            file.seek(0)
            df = pd.read_excel(file, header=0)
    return df


def extract_writing_dna(df: pd.DataFrame, max_samples: int = 5) -> str:
    name_col = find_col(df, 'name')
    desc_col = find_col(df, 'description')
    cat_col = find_col(df, 'category')
    brand_col = get_brand_col(df)
    price_col = find_col(df, 'price')

    samples = []
    if name_col and desc_col:
        for _, row in df.iterrows():
            name = str(row.get(name_col, ''))
            desc = str(row.get(desc_col, ''))
            if (not is_tester(name) and len(desc) > 200 and '<' in desc):
                samples.append({
                    'name': name,
                    'brand': str(row.get(brand_col, '')) if brand_col else '',
                    'category': str(row.get(cat_col, '')) if cat_col else '',
                    'price': row.get(price_col, 0) if price_col else 0,
                    'description_sample': desc[:800],
                })
            if len(samples) >= max_samples:
                break

    all_categories = []
    if cat_col:
        all_categories = sorted(df[cat_col].dropna().astype(str).unique().tolist())

    dna = "## أسلوب الكتابة المُتَّبع في متجر مهووس (تعلّم منه ولا تخرج عنه)\n\n"
    dna += "### التصنيفات المتاحة (استخدمها حرفياً):\n"
    dna += "\n".join(f"- {c}" for c in all_categories) + "\n\n"
    dna += "### أمثلة فعلية من أوصاف المنتجات (انسخ الأسلوب والتنسيق):\n"
    for i, s in enumerate(samples, 1):
        dna += (
            f"\n--- مثال {i} ---\n"
            f"الاسم: {s['name']}\n"
            f"الماركة: {s['brand']}\n"
            f"التصنيف: {s['category']}\n"
            f"السعر: {s['price']} ريال\n"
            f"مقطع من الوصف:\n{s['description_sample']}\n---\n"
        )
    return dna


def extract_json(text: str) -> dict:
    text = re.sub(r'^\s*```(?:json)?\s*\n?', '', text, flags=re.MULTILINE)
    text = re.sub(r'\n?\s*```\s*$',          '', text, flags=re.MULTILINE)
    text = text.strip()

    start = text.find('{')
    end   = text.rfind('}')
    if start == -1 or end == -1:
        snippet = (text[:300] + '...') if len(text) > 300 else text
        raise ValueError(f"لم يُرجع Gemini JSON صالحاً ({len(text)} حرف): {snippet!r}")
    raw = text[start:end + 1]

    try:
        return json.loads(raw, strict=False)
    except (json.JSONDecodeError, ValueError):
        pass

    raw = re.sub(r'\\u(?![0-9a-fA-F]{4})', r'\\\\u', raw)
    raw = re.sub(r'\\(?!["\\/bfnrtu])', r'\\\\', raw)

    result, in_string, escape_next = [], False, False
    for ch in raw:
        if escape_next:
            result.append(ch); escape_next = False
        elif ch == '\\' and in_string:
            result.append(ch); escape_next = True
        elif ch == '"':
            in_string = not in_string; result.append(ch)
        elif in_string and ch == '\n':
            result.append('\\n')
        elif in_string and ch == '\r':
            result.append('\\r')
        elif in_string and ch == '\t':
            result.append('\\t')
        else:
            result.append(ch)
    cleaned = ''.join(result)
    cleaned = re.sub(r',(\s*[}\]])', r'\1', cleaned)

    try:
        return json.loads(cleaned, strict=False)
    except (json.JSONDecodeError, ValueError):
        pass

    cleaned2 = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', cleaned)
    try:
        return json.loads(cleaned2, strict=False)
    except (json.JSONDecodeError, ValueError):
        pass

    try:
        from json_repair import repair_json
        repaired = repair_json(cleaned2, return_objects=True)
        if isinstance(repaired, dict):
            return repaired
    except Exception:
        pass

    snippet = (cleaned2[:400] + '...') if len(cleaned2) > 400 else cleaned2
    raise ValueError(f"فشل تحليل JSON بعد كل المحاولات. المقتطف: {snippet!r}")


# ═════════════════════════════════════════════════════════════════════════════
#  ⭐ اختبار مبكر لمفتاح Gemini — يكشف 403 وأخطاء المفتاح قبل بدء المعالجة
# ═════════════════════════════════════════════════════════════════════════════

def test_gemini_key(api_key: str, model_name: str = 'gemini-2.5-flash') -> tuple[bool, str]:
    """يختبر مفتاح Gemini بطلب بسيط (5 توكن) للكشف المبكر عن المشاكل.

    يعيد (نجاح: bool، رسالة عربية للعرض في الواجهة).
    يكتشف:
      • 403 PERMISSION_DENIED (مشروع محجوب على Cloud)
      • 401 / مفتاح غير صحيح
      • مفاتيح مُبلَّغ عنها كمُسرَّبة
      • تجاوز الحصة (429)
      • مشاكل اتصال بالإنترنت
    """
    if not api_key or not api_key.strip():
        return False, "⚠️ لم يُدخل مفتاح API"
    if not api_key.startswith('AIza'):
        return False, (
            "⚠️ صيغة المفتاح غير صحيحة — مفاتيح Gemini تبدأ بـ 'AIza'. "
            "تأكد من نسخ المفتاح كاملاً من Google AI Studio."
        )
    try:
        client = genai.Client(api_key=api_key)
        config = genai_types.GenerateContentConfig(
            temperature=0.0,
            max_output_tokens=5,
        )
        resp = client.models.generate_content(
            model=model_name,
            contents='test',
            config=config,
        )
        if resp.text:
            return True, f"✅ المفتاح يعمل بنجاح مع {model_name}"
        return False, "⚠️ المفتاح صحيح لكن الرد فارغ — جرّب موديلاً آخر"
    except Exception as e:
        err = str(e).lower()
        if ('403' in err and 'denied access' in err) or 'permission_denied' in err:
            return False, (
                "❌ **المشروع محجوب من Google** (403 PERMISSION_DENIED)\n\n"
                "🔴 هذا الخطأ على مستوى حساب Google وليس في التطبيق.\n\n"
                "**الأسباب المحتملة:**\n"
                "• المفتاح تسرّب على GitHub وحُجب تلقائياً\n"
                "• المشروع محجوب على Google Cloud\n"
                "• منطقتك تتطلب تفعيل Cloud Billing\n\n"
                "**الحل (مرتّب حسب الأولوية):**\n"
                "1. اذهب إلى https://aistudio.google.com/apikey\n"
                "2. احذف المفتاح القديم (أيقونة سلة المهملات)\n"
                "3. اضغط **Create API Key** → اختر **Create in new project**\n"
                "4. استخدم المفتاح الجديد هنا\n\n"
                "إن استمرّ الخطأ: فعّل Billing من https://console.cloud.google.com/billing"
            )
        if '401' in err or 'unauthenticated' in err or 'invalid api' in err or 'api_key_invalid' in err:
            return False, (
                "❌ **المفتاح غير صحيح أو منتهٍ**\n\n"
                "أنشئ مفتاحاً جديداً من https://aistudio.google.com/apikey "
                "وتأكد من نسخه كاملاً."
            )
        if 'leaked' in err or 'reported as leaked' in err:
            return False, (
                "❌ **المفتاح مُبلَّغ عنه كمُسرَّب**\n\n"
                "Google كشفه على GitHub أو موقع عام وحجبه. "
                "أنشئ مفتاحاً جديداً وتجنّب رفعه لـ GitHub أو مشاركته علناً."
            )
        if '429' in err or 'quota' in err or 'resource_exhausted' in err:
            return False, (
                "⚠️ **تجاوزت حصة الاستخدام**\n\n"
                "جرّب gemini-2.5-flash-lite (حصة أكبر بـ 4 مرات) أو "
                "انتظر حتى منتصف الليل بتوقيت المحيط الهادئ."
            )
        if 'connection' in err or 'timeout' in err or 'network' in err:
            return False, f"⚠️ مشكلة اتصال بالإنترنت: {e}"
        return False, f"❌ خطأ غير متوقع: {e}"


# ═════════════════════════════════════════════════════════════════════════════
#  محرّك التطبيع والمقارنة التقريبية (Fuzzy Matching)
# ═════════════════════════════════════════════════════════════════════════════

_AR_VOWELS = set('اويىآأإٱةؤئءَُِّْ')

_AR_CONSONANTS = {
    'ب': 'b', 'ت': 't', 'ث': 't',
    'ج': 'g', 'ح': 'h', 'خ': 'k',
    'د': 'd', 'ذ': 'd',
    'ر': 'r', 'ز': 'z',
    'س': 's', 'ش': 's', 'ص': 's', 'ض': 'd',
    'ط': 't', 'ظ': 'z',
    'ع': '', 'غ': 'g',
    'ف': 'f', 'ق': 'k', 'ك': 'k',
    'ل': 'l', 'م': 'm', 'ن': 'n',
    'ه': 'h',
}

_LATIN_DIGRAPHS = [
    ('ph', 'f'), ('th', 't'), ('ch', 's'), ('sh', 's'),
    ('kh', 'k'), ('gh', 'g'), ('qu', 'k'),
]

_LATIN_REPLACE = [
    ('v', 'f'), ('j', 'g'), ('x', 'k'), ('q', 'k'),
]

_LATIN_VOWELS_DROP = set('aeiouy')

_LAT_PHRASES = [
    ('eau de parfum', 'edp'),
    ('eau de toilette', 'edt'),
    ('eau de cologne', 'edc'),
]

_JUNK_WORDS = {
    'للرجال', 'للنساء', 'رجالي', 'نسائي', 'النسائي', 'الرجالي',
    'عطر', 'العطر', 'تستر', 'تيستر', 'تستير', 'tester', 'testr',
    'مل', 'ml', 'بخاخ', 'spray', 'للجنسين', 'unisex',
    'قديم', 'جديد', 'الجديد', 'القديم', 'النسخه', 'النسخة',
    'edp', 'edt', 'edc',
    'pour', 'homme', 'femme', 'for', 'men', 'women', 'man', 'woman',
    'eau', 'de', 'la', 'le', 'دي', 'دو', 'لو',
    'الاصلي', 'original', 'authentic', 'اصلي',
}

_SIZE_RX = re.compile(
    r'(\d+(?:[.,]\d+)?)\s*(?:ml|مل|مللي|مللتر|milliliter)\b',
    re.IGNORECASE,
)


def _extract_size_ml(name: str) -> int:
    """يُرجع الحجم بالـ مل (0 إذا لم يُذكر)."""
    if not name:
        return 0
    m = _SIZE_RX.search(str(name))
    if not m:
        return 0
    try:
        return int(float(m.group(1).replace(',', '.')))
    except (ValueError, TypeError):
        return 0


def _normalize_perfume_name(name: str) -> str:
    """يُنتج «هيكلاً صوتياً» قابلاً للمقارنة عبر اللغات."""
    if not name:
        return ''
    s = str(name).lower().strip()
    s = re.sub(r'[\u064b-\u065f\u0670]', '', s)
    s = (s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا').replace('ٱ', 'ا')
           .replace('ى', 'ي').replace('ة', 'ه'))
    s = _SIZE_RX.sub(' ', s)
    s = re.sub(r'\b\d+\b', ' ', s)
    for phrase, repl in _LAT_PHRASES:
        s = s.replace(phrase, repl)
    for pat, repl in _LATIN_DIGRAPHS:
        s = s.replace(pat, repl)
    for pat, repl in _LATIN_REPLACE:
        s = s.replace(pat, repl)

    out_chars = []
    for ch in s:
        if ch in _AR_VOWELS:
            continue
        elif ch in _AR_CONSONANTS:
            out_chars.append(_AR_CONSONANTS[ch])
        elif ch.isalnum() or ch.isspace():
            out_chars.append(ch)
    s = ''.join(out_chars)

    tokens = s.split()
    cleaned = []
    for tok in tokens:
        if tok.startswith('ال') and len(tok) > 3:
            tok = tok[2:]
        if not tok or tok in _JUNK_WORDS or len(tok) < 2:
            continue
        consonants = ''.join(c for c in tok if c not in _LATIN_VOWELS_DROP and c != 'h')
        if consonants:
            cleaned.append(consonants)

    skeleton = ''.join(cleaned)
    skeleton = re.sub(r'(.)\1+', r'\1', skeleton)
    return skeleton


def _name_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if len(a) >= 4 and len(b) >= 4 and (a in b or b in a):
        ratio_len = min(len(a), len(b)) / max(len(a), len(b))
        if ratio_len >= 0.7:
            return 0.95
    return SequenceMatcher(None, a, b).ratio()


def filter_duplicates(result: dict, existing_products: list,
                      similarity_threshold: float = 0.85) -> dict:
    existing_keys = []
    for p in existing_products:
        nm = p.get('name', '')
        sk = _normalize_perfume_name(nm)
        sz = _extract_size_ml(nm)
        if sk:
            existing_keys.append((sk, sz))

    def _matches_existing(skeleton, size):
        if not skeleton:
            return True
        for ek, es in existing_keys:
            if size and es and size != es:
                continue
            if _name_similarity(skeleton, ek) >= similarity_threshold:
                return True
        return False

    def _is_internal_dup(skeleton, size, seen):
        if not skeleton:
            return True
        for sk, sz in seen:
            if size and sz and size != sz:
                continue
            if _name_similarity(skeleton, sk) >= similarity_threshold:
                return True
        return False

    # ⚠️ ملاحظة: التساتر تُدَقّق على عدم تكرار base_product_id فقط،
    # ولا تُدقَّق ضد existing_products لأنها بطبيعتها مشتقة من العطور الأساسية.
    testers = result.get('testers_to_add')
    if isinstance(testers, list):
        seen_base_ids: set = set()
        seen_keys: list = []
        deduped = []
        for t in testers:
            if not isinstance(t, dict):
                continue
            base_id = str(t.get('base_product_id', '') or '').strip()
            nm = t.get('name', '')
            sk = _normalize_perfume_name(nm)
            sz_from_name = _extract_size_ml(nm)
            try:
                sz_from_field = int(float(t.get('size_ml') or 0))
            except (ValueError, TypeError):
                sz_from_field = 0
            sz = sz_from_name or sz_from_field

            if base_id and base_id in seen_base_ids:
                continue
            if _is_internal_dup(sk, sz, seen_keys):
                continue

            # تأكد من وجود الحقول الجديدة
            t.setdefault('competitor_price', 0)
            t.setdefault('tester_available_in_market', False)
            t.setdefault('source_store', '')

            # وحّد صيغة الاسم: «تستر» في البداية وليس النهاية
            # (Gemini أحياناً تُلصقها في النهاية بصيغة خاطئة لسلة)
            if nm:
                normalized_name = build_tester_name(nm)
                if normalized_name != nm:
                    t['name'] = normalized_name

            if base_id:
                seen_base_ids.add(base_id)
            seen_keys.append((sk, sz))
            deduped.append(t)
        result['testers_to_add'] = deduped

    missing = result.get('missing_products')
    if isinstance(missing, list):
        seen_keys = []
        deduped = []
        for m in missing:
            if not isinstance(m, dict):
                continue
            nm = m.get('name', '')
            sk = _normalize_perfume_name(nm)
            sz_from_name = _extract_size_ml(nm)
            try:
                sz_from_field = int(float(m.get('size_ml') or 0))
            except (ValueError, TypeError):
                sz_from_field = 0
            sz = sz_from_name or sz_from_field

            if _is_internal_dup(sk, sz, seen_keys):
                continue
            if _matches_existing(sk, sz):
                continue

            m.setdefault('image_url_1', '')
            m.setdefault('image_url_2', '')
            if m.get('image_url_2') is None:
                m['image_url_2'] = ''

            seen_keys.append((sk, sz))
            deduped.append(m)
        result['missing_products'] = deduped

    upd = result.get('testers_updated')
    if isinstance(upd, list):
        seen_keys = []
        kept = []
        for t in upd:
            if not isinstance(t, dict):
                continue
            sk = _normalize_perfume_name(t.get('name', ''))
            sz = _extract_size_ml(t.get('name', ''))
            if _is_internal_dup(sk, sz, seen_keys):
                continue
            if t.get('is_new') and _matches_existing(sk, sz):
                continue
            seen_keys.append((sk, sz))
            kept.append(t)
        result['testers_updated'] = kept

    return result


# ═════════════════════════════════════════════════════════════════════════════
#  ⭐ شبكة الأمان: ضمان وجود تستر لكل عطر أساسي بدون تستر
# ═════════════════════════════════════════════════════════════════════════════

# regex لإزالة كلمات «الضوضاء» من الأسماء قبل التطبيع.
# ضرورية لأن `_normalize_perfume_name` يفحص `_JUNK_WORDS` *بعد* التحويل
# للحروف اللاتينية، لكن قائمة الكلمات تحتوي الأصل العربي — فلا يتطابق.
# نُزيلها هنا قبل التطبيع لتفادي اختلاف الهياكل بسبب كلمات حشو مثل
# "عطر" و "تستر" و "للرجال" إلخ.
_NOISE_STRIP_RX = re.compile(
    r'\b(?:tester|testr|test)\b'
    r'|تستر|تيستر|تستير'
    r'|العطر|عطر'
    r'|الرجالي|الرجالى|للرجال|رجالي|رجالى'
    r'|النسائي|النسائى|للنساء|نسائي|نسائى'
    r'|للجنسين|unisex'
    r'|الاصلي|الأصلي|اصلي|أصلي|original|authentic',
    re.IGNORECASE,
)

# alias للتوافق الخلفي
_TESTER_STRIP_RX = _NOISE_STRIP_RX


def _strip_tester_keyword(name: str) -> str:
    """يُزيل كلمات الحشو الشائعة (تستر/عطر/للرجال/الأصلي…) من الاسم قبل
    التطبيع. مهم جداً لأن `_normalize_perfume_name` لا يستطيع إزالة هذه
    الكلمات العربية بنفسه (لأنه يفحص قائمة `_JUNK_WORDS` بعد تحويل الأحرف
    إلى لاتينية، فلا يتطابق الأصل العربي).
    """
    if not name:
        return ''
    return _NOISE_STRIP_RX.sub(' ', str(name))


def ensure_all_testers_added(result: dict, products_payload: list) -> dict:
    """شبكة أمان نهائية: تضمن وجود تستر لكل عطر أساسي ليس له تستر،
    حتى لو فاتت Gemini أحدها. تستخدم قاعدة التسعير الداخلية إذا لم يكن
    هناك سعر منافس مرجعي.

    - `result`: قاموس النتائج المُجمّعة (بعد merge_batch_results)
    - `products_payload`: قائمة منتجات الماركة الكاملة
    """
    if not isinstance(result, dict):
        return result
    if not products_payload:
        return result

    # افصل العطور الأساسية عن التساتر الموجودة في القائمة الأصلية
    # ⚠️ نستثني الأطقم/المجموعات: لا نضيف تستر للأطقم.
    base_perfumes = [
        p for p in products_payload
        if isinstance(p, dict)
        and not is_tester(p.get('name', ''))
        and not is_set(p.get('name', ''))
    ]
    existing_tester_products = [
        p for p in products_payload
        if isinstance(p, dict) and is_tester(p.get('name', ''))
    ]

    # ⭐ قاعدة الحجم الأكبر: إذا كان نفس العطر متوفراً بأحجام متعددة،
    # نختار فقط الحجم الأكبر للحصول على تستر (لا تستر للأحجام الصغيرة).
    # نُجمّع العطور حسب هيكل الاسم بدون الحجم.
    size_groups: dict = {}  # {skeleton_no_size: [(idx, size, price, name), ...]}
    for bp in base_perfumes:
        bp_name = (bp.get('name', '') or '').strip()
        if not bp_name:
            continue
        sk_no_size = _normalize_perfume_name(_strip_tester_keyword(bp_name))
        sz = _extract_size_for_grouping(bp_name)
        size_groups.setdefault(sk_no_size, []).append(bp)

    # في كل مجموعة، احتفظ فقط بالحجم الأكبر
    largest_in_group = set()  # ids of base perfumes that should get a tester
    skipped_for_size = []
    for sk, items in size_groups.items():
        if not sk:  # مفقود الهيكل — تخطه
            continue
        if len(items) == 1:
            largest_in_group.add(id(items[0]))
            continue
        # احسب الأحجام واختر الأكبر
        sized = [(it, _extract_size_for_grouping(it.get('name', '') or '')) for it in items]
        max_size = max((s for _, s in sized), default=0)
        if max_size == 0:
            # لا أحجام معروفة — اختر الأول وتخطّى الباقي
            largest_in_group.add(id(items[0]))
            for it in items[1:]:
                skipped_for_size.append(it.get('name', ''))
            continue
        # احتفظ بكل من له الحجم الأكبر، تخطّ الأصغر
        for it, sz in sized:
            if sz == max_size:
                largest_in_group.add(id(it))
            else:
                skipped_for_size.append(it.get('name', ''))

    # مجموعة هياكل أسماء المنتجات التي لها تستر بالفعل في القائمة الأصلية
    has_tester_skeletons: set = set()
    for tp in existing_tester_products:
        raw = tp.get('name', '') or ''
        stripped = _strip_tester_keyword(raw)
        sk = _normalize_perfume_name(stripped)
        if sk:
            has_tester_skeletons.add(sk)

    # base_product_ids التي اقترحت لها Gemini تستر بالفعل
    existing_testers = result.get('testers_to_add', []) or []
    if not isinstance(existing_testers, list):
        existing_testers = []
    existing_base_ids = {
        str(t.get('base_product_id', '') or '').strip()
        for t in existing_testers
        if isinstance(t, dict) and str(t.get('base_product_id', '') or '').strip()
    }

    # هياكل أسماء التساتر التي اقترحتها Gemini (للحماية من التكرار حتى عند
    # غياب base_product_id)
    existing_tester_skeletons: set = set()
    for t in existing_testers:
        if not isinstance(t, dict):
            continue
        raw = t.get('name', '') or ''
        stripped = _strip_tester_keyword(raw)
        sk = _normalize_perfume_name(stripped)
        if sk:
            existing_tester_skeletons.add(sk)

    auto_added_names = []

    for bp in base_perfumes:
        bp_id = str(bp.get('id', '') or '').strip()
        bp_name = (bp.get('name', '') or '').strip()
        if not bp_name:
            continue

        # ⭐ تخطّ إذا لم يكن هذا الحجم الأكبر لمجموعته
        if id(bp) not in largest_in_group:
            continue

        bp_sk = _normalize_perfume_name(_strip_tester_keyword(bp_name))
        try:
            bp_price = float(bp.get('price', 0) or 0)
        except (TypeError, ValueError):
            bp_price = 0.0

        # 🖼️ احتفظ بقائمة الصور كاملة (مفصولة بفواصل) — Salla يدعم صور متعددة
        raw_img = (bp.get('image_url') or '').strip()
        # نظّف ونزّل الفواصل
        if raw_img:
            img_list = [u.strip() for u in raw_img.split(',') if u.strip()]
            bp_img = ','.join(img_list)
        else:
            bp_img = ''

        # تخطّ إذا اقترحت Gemini له تستر بالفعل (عبر base_product_id)
        if bp_id and bp_id in existing_base_ids:
            continue
        # تخطّ إذا اقترحت Gemini له تستر بالفعل (عبر هيكل الاسم)
        if bp_sk and bp_sk in existing_tester_skeletons:
            continue
        # تخطّ إذا كان له تستر فعلاً في القائمة الأصلية
        if bp_sk and bp_sk in has_tester_skeletons:
            continue

        # أضف تستراً تلقائياً بقاعدة التسعير الداخلية
        tester_price = calc_tester_price(bp_price)
        size_ml = _extract_size_for_grouping(bp_name) or 100
        bp_brand = (bp.get('brand') or '').strip()  # قد يكون موجوداً في products_payload
        bp_category = (bp.get('category') or '').strip()

        # 📝 ابنِ وصفاً مُملوءاً بالكامل (بدون placeholders) كإفتراضي.
        # `enrich_auto_added_testers` لاحقاً تستبدل النوتات والعائلة بمعلومات من Gemini.
        clean_n = clean_perfume_display_name(bp_name)
        clean_b = clean_brand_display_name(bp_brand)
        gender_disp = detect_gender(bp_name, bp_category)
        complete_desc = fill_tester_template_complete(
            brand_name=bp_brand or '',  # يُحدَّث في build_output_excel
            perfume_name=bp_name,
            size_ml=size_ml,
            base_category=bp_category,
            enrichment=None,  # سيُعبَّأ لاحقاً
        )

        auto_tester = {
            'base_product_id': bp_id,
            'name': build_tester_name(bp_name),
            'size_ml': size_ml,
            'original_price': bp_price,
            'new_price': tester_price,
            'competitor_price': 0,
            'image_url': bp_img,  # ← قائمة كاملة بالصور
            'source_store': '',
            'tester_available_in_market': False,
            'new_description': complete_desc,  # ← مُملوء بدون placeholders
            'seo_title': build_tester_name(bp_name)[:60],
            'seo_description': (
                f"تستر {clean_n} من {clean_b} الأصلي 100% — نفس السائل "
                f"والثبات والفوحان للإصدار المغلف بسعر استثنائي."
                if clean_b else
                f"تستر {clean_n} الأصلي 100% — نفس السائل والثبات والفوحان "
                f"للإصدار المغلف بسعر استثنائي."
            )[:155],
            '_auto_added': True,
            '_base_category': bp_category,  # للاستخدام في تصنيف التستر لاحقاً
            '_base_perfume_name': bp_name,
        }
        existing_testers.append(auto_tester)
        if bp_id:
            existing_base_ids.add(bp_id)
        if bp_sk:
            existing_tester_skeletons.add(bp_sk)
        auto_added_names.append(bp_name)

    result['testers_to_add'] = existing_testers
    if auto_added_names:
        # سجل ما أُضيف تلقائياً للمراجعة لاحقاً
        prev = result.get('_auto_added_testers') or []
        if not isinstance(prev, list):
            prev = []
        result['_auto_added_testers'] = prev + auto_added_names
    if skipped_for_size:
        # سجل ما تم تخطيه بسبب وجود حجم أكبر
        result['_skipped_smaller_sizes'] = skipped_for_size

    return result


def merge_batch_results(accum: dict, new: dict) -> dict:
    if not accum:
        accum = {
            'brand': new.get('brand', ''),
            'products_updated': [],
            'testers_to_add': [],
            'orphan_testers': [],
            'missing_products': []
        }
    existing_ids = {str(t.get('base_product_id', '')) for t in accum.get('testers_to_add', [])}
    for t in new.get('testers_to_add', []):
        bid = str(t.get('base_product_id', ''))
        if bid and bid not in existing_ids:
            accum['testers_to_add'].append(t)
            existing_ids.add(bid)
    existing_norms = {_normalize_perfume_name(m.get('name', '')) for m in accum.get('missing_products', [])}
    for m in new.get('missing_products', []):
        norm = _normalize_perfume_name(m.get('name', ''))
        if norm and norm not in existing_norms:
            accum['missing_products'].append(m)
            existing_norms.add(norm)
    if 'products_updated' in new:
        accum['products_updated'].extend(new['products_updated'])
    if 'orphan_testers' in new:
        accum['orphan_testers'].extend(new['orphan_testers'])
    return accum


# ═════════════════════════════════════════════════════════════════════════════
#  call_gemini_brand — إعادة محاولة ذكية، البحث محفوظ دائماً
# ═════════════════════════════════════════════════════════════════════════════

def call_gemini_brand(
    brand_name: str,
    products: list,
    full_brand_products: list,
    api_key: str,
    writing_dna: str,
    model_name: str = 'gemini-2.5-flash',
    use_grounding: bool = True,
    batch_index: int = 0,
    total_batches: int = 1,
    progress_cb=None,
    max_retries: int = 4,
    base_backoff_seconds: float = 6.0,
) -> dict:
    """يستدعي Gemini لمعالجة دفعة من ماركة واحدة.

    - إعادة محاولة بانتظار تصاعدي للأخطاء العابرة (6→18→54→162ث)
    - رفع فوري للأخطاء الصلبة (مفتاح API، فوترة، أمان)
    - عند استنفاد المحاولات → RuntimeError واضح
    - البحث (Grounding) محفوظ دائماً — لا يُعطّل عند الفشل
    """
    client = genai.Client(api_key=api_key)

    system_instruction = SYSTEM_INSTRUCTION_TEMPLATE.format(
        writing_dna=writing_dna,
        HTML_TEMPLATE_NEW=HTML_TEMPLATE_NEW,
        HTML_TEMPLATE_TESTER=HTML_TEMPLATE_TESTER,
        competitors_json=json.dumps(COMPETITOR_STORES, ensure_ascii=False),
    )

    base_perfumes = [p for p in full_brand_products if not is_tester(p.get('name', ''))]
    tester_products = [p for p in full_brand_products if is_tester(p.get('name', ''))]

    base_catalog_json = json.dumps(
        [{'id': p['id'], 'name': p['name'], 'price': p.get('price', 0),
          'image_url': (p.get('image_url') or '').split(',')[0].strip()}
         for p in base_perfumes],
        ensure_ascii=False, indent=2
    )

    tester_catalog_json = json.dumps(
        [{'id': p['id'], 'name': p['name'], 'price': p.get('price', 0)}
         for p in tester_products],
        ensure_ascii=False, indent=2
    )

    prompt = f"""أنت تعالج ماركة "{brand_name}" — الدفعة {batch_index + 1} من {total_batches}.

## قائمة العطور الأساسية لدينا (غير التساتر) — {len(base_perfumes)} عطر:
{base_catalog_json}

## التساتر الموجودة لدينا حالياً — {len(tester_products)} تستر:
{tester_catalog_json}

---

## المهمة 1: إضافة التستر لكل عطر أساسي بلا تستر (إلزامي)
لكل عطر في قائمة "العطور الأساسية":
1. تحقق: هل يوجد في "التساتر الموجودة" تستر يحمل نفس الاسم (بأي صيغة عربية أو إنجليزية)؟
   - إذا نعم → لا تفعل شيئاً (تخطّ).
   - إذا لا → تابع الخطوة 2 (إلزامي):

2. ابحث في المتاجر السعودية للحصول على **سعر مرجعي** للتستر، ثم **أضف التستر دائماً** في `testers_to_add` مع:
     * `name`: اسم العطر + " تستر"
     * `size_ml`: نفس حجم العطر الأساسي ما لم يُذكر حجم تستر مختلف عند منافس
     * `base_product_id`: id العطر الأساسي من قائمتنا (إلزامي ودقيق)
     * `image_url`: انسخه حرفياً من حقل image_url للعطر الأساسي (الصورة الأولى فقط) — لا تبحث عن صورة جديدة
     * `original_price`: سعر العطر الأساسي من قائمتنا
     * `new_price`: مطبقاً قاعدة التسعير الداخلية
       (أقل من 1000 → ناقص 70؛ 1000 فأكثر → ناقص 150)
     * `competitor_price`: سعر التستر عند المنافس (مرجعي فقط)، أو 0 إذا لم تجد
     * `tester_available_in_market`: true إذا وجدته عند منافس سعودي، false إذا لم تجد
     * `source_store`: اسم المتجر السعودي إن وُجد، أو "" إذا لم تجد
     * `new_description`: قالب التستر مكتملاً
     * `seo_title` و `seo_description`: قصيران ومحسّنان

   - 🔴 **لا تتجاوز هذه الخطوة أبداً** — كل عطر بلا تستر **يجب** أن يحصل على تستر، حتى لو لم تجد التستر عند أي منافس. ضع `competitor_price=0` و `tester_available_in_market=false` و `source_store=""` وأضفه.

## المهمة 2: التساتر التي ليس لها عطر أساسي (Orphan Testers)
لكل تستر في "التساتر الموجودة":
1. تحقق: هل يوجد في "العطور الأساسية" منتج بنفس الاسم (بدون كلمة تستر)؟
   - إذا نعم → تخطّ.
   - إذا لا → أضف المنتج الأساسي في `missing_products` كعطر جديد يجب إضافته:
     * ابحث عن صورة الزجاجة من المتاجر السعودية أو الموقع الرسمي للماركة
     * اكتب وصفاً بقالب العطور الجديدة
     * واذكر التستر اليتيم في `orphan_testers` للعرض في الواجهة.

## المهمة 3: المنتجات الناقصة عند المنافسين
قارن قائمتنا الكاملة ({len(full_brand_products)} منتج) بما يبيعه المنافسون السعوديون لماركة "{brand_name}".
- 🔴 الأولوية القصوى: إصدارات 2024 و2025 و2026 — ابحث بالاسم الصريح مثل "Million Gold" و"Phantom Intense" و"Phantom Elixir" وما صدر حديثاً
- ركّز على: الأكثر مبيعاً، الأحجام الشائعة (50مل، 100مل، 200مل)، الإصدارات الجديدة
- لا تكتفِ بعطور 2022 و2023 — ابحث صراحةً عن "{brand_name} new release 2024 2025 2026" في المتاجر السعودية
- اقترح فقط المنتجات المتوفرة للشراء الآن مع ذكر المتجر المصدر
- لكل منتج مقترح: اكتب وصفاً كاملاً بقالب العطور الجديدة

## ⚠️ استراتيجية البحث الإلزامية — لا تتجاوزها
لكل عطر تبحث عنه، **افتح هذه المتاجر بالترتيب** وابحث فيها فعلياً:
1. https://www.noon.com/saudi-ar/ — ابحث: "[اسم العطر] {brand_name} tester"
2. https://en.ounass.com/saudi-arabia/ — ابحث: "{brand_name} perfume tester"
3. https://www.goldenscent.com/ — ابحث مباشرةً باسم العطر + tester
4. https://niceonesa.com/ — ابحث مباشرةً باسم العطر + tester
5. https://www.amazon.sa/ — ابحث: "{brand_name} [perfume name] tester"

**قاعدة صارمة:** لا تكتفِ بمتجرَين. إذا لم تجد في الأول، انتقل للثاني والثالث.
**لكن: حتى لو لم تجد التستر في أي مكان، أضفه إلى testers_to_add بـ competitor_price=0.**

## ⚠️ تحذير نهائي قبل الإخراج
- قبل إرجاع JSON، راجع المصفوفات وتأكد:
  1. **كل عطر أساسي بلا تستر له عنصر مقابل في testers_to_add** (لا استثناءات).
  2. لا يوجد base_product_id مكرر داخل testers_to_add (واحد فقط لكل عطر).
  3. لا يوجد منتج مكرر داخل missing_products بأي صيغة (عربي/إنجليزي/أحجام مختلفة بنفس المنتج).
  4. كل عنصر في testers_to_add يحتوي `competitor_price` و `tester_available_in_market` (حتى لو 0 و false).
  5. كل عنصر في missing_products يحتوي image_url_1 و image_url_2 (الثاني قد يكون "" لكنه موجود).

**أعد JSON صارم فقط:**

{{
  "brand": "{brand_name}",
  "batch_index": {batch_index},
  "products_updated": [],
  "testers_to_add": [
    {{
      "base_product_id": "id من قائمتنا",
      "name": "اسم العطر تستر",
      "size_ml": 100,
      "original_price": 0,
      "new_price": 0,
      "competitor_price": 0,
      "image_url": "منسوخ حرفياً من العطر الأساسي",
      "source_store": "اسم المتجر السعودي أو \\"\\"",
      "tester_available_in_market": false,
      "new_description": "<p>...</p>",
      "seo_title": "عنوان أقل من 60 حرف",
      "seo_description": "وصف أقل من 155 حرف"
    }}
  ],
  "orphan_testers": [
    {{
      "tester_product_id": "id التستر الذي ليس له أساسي",
      "tester_name": "اسم التستر",
      "suggested_base_name": "اسم العطر الأساسي المقترح إضافته"
    }}
  ],
  "missing_products": [
    {{
      "name": "اسم العطر الكامل",
      "type": "عطر مفرد",
      "category": "من التصنيفات المتاحة حرفياً",
      "price": 0,
      "size_ml": 100,
      "concentration": "EDP",
      "gender": "رجالي",
      "description": "<p>...</p>",
      "brand": "{brand_name}",
      "is_tester": false,
      "is_bestseller": true,
      "source_store": "اسم المتجر السعودي",
      "image_url_1": "زجاجة بخلفية بيضاء",
      "image_url_2": "زجاجة + كرتون بخلفية بيضاء",
      "seo_title": "عنوان أقل من 60 حرف",
      "seo_description": "وصف أقل من 155 حرف"
    }}
  ]
}}"""

    config_kwargs = dict(
        system_instruction=system_instruction,
        temperature=0.0,
        max_output_tokens=65536,
    )
    if use_grounding:
        config_kwargs['tools'] = [genai_types.Tool(google_search=genai_types.GoogleSearch())]
    else:
        config_kwargs['response_mime_type'] = 'application/json'

    config = genai_types.GenerateContentConfig(**config_kwargs)

    last_err = None

    for attempt in range(1, max_retries + 1):
        try:
            stream = client.models.generate_content_stream(
                model=model_name,
                contents=prompt,
                config=config,
            )

            text = ''
            last_chunk = None
            for chunk in stream:
                last_chunk = chunk
                t = ''
                try:
                    t = chunk.text or ''
                except Exception:
                    for cand in getattr(chunk, 'candidates', []) or []:
                        content = getattr(cand, 'content', None)
                        if not content:
                            continue
                        for part in getattr(content, 'parts', []) or []:
                            t += getattr(part, 'text', '') or ''
                if t:
                    text += t
                    if progress_cb:
                        try:
                            progress_cb(len(text))
                        except Exception:
                            pass

            finish = ''
            safety = ''
            try:
                finish = str(last_chunk.candidates[0].finish_reason) if last_chunk else ''
                safety = (
                    str(getattr(last_chunk.candidates[0], 'safety_ratings', ''))[:200]
                    if last_chunk else ''
                )
            except Exception:
                pass

            if not text.strip():
                if 'SAFETY' in finish:
                    raise RuntimeError(f"حُجبت الاستجابة بفلتر الأمان (safety={safety})")
                hint = ' — قلّل BATCH_SIZE.' if 'MAX_TOKENS' in finish else ''
                raise ValueError(f"Gemini أعاد رداً فارغاً (finish_reason={finish}){hint}")

            try:
                return extract_json(text)
            except (ValueError, json.JSONDecodeError) as e:
                raise ValueError(f"{e} | finish_reason={finish}") from e

        except Exception as e:
            last_err = e
            err_low = str(e).lower()

            HARD_ERRORS = (
                'api key', 'invalid api', 'api_key', 'unauthenticated',
                'authentication', 'permission_denied', 'permission denied',
                'billing', 'safety', 'حُجبت', 'blocked',
            )
            if any(x in err_low for x in HARD_ERRORS):
                raise RuntimeError(f"❌ خطأ صلب في Gemini (لا يُعاد المحاولة): {e}") from e

            # ⚠️ إذا تجاوزت الحصة اليومية (RPD)، لا فائدة من إعادة المحاولة —
            # الانتظار حتى منتصف الليل بتوقيت المحيط الهادئ ضروري.
            DAILY_QUOTA_INDICATORS = (
                'exceeded your current quota', 'exceeded your quota',
                'requests per day', 'rpd', 'daily quota', 'per_day',
                'generate_requests_per_model_per_day',
            )
            if '429' in err_low and any(x in err_low for x in DAILY_QUOTA_INDICATORS):
                raise RuntimeError(
                    f"❌ تجاوزت الحصة اليومية لـ Gemini (RPD).\n"
                    f"الحلول: (1) انتظر حتى 10:00 صباحاً بتوقيت السعودية لإعادة الحصة. "
                    f"(2) جرّب نموذج 'gemini-2.5-flash-lite' (حصة أكبر بـ 4 مرات). "
                    f"(3) فعّل Cloud Billing على حسابك (تيار 1) — يرفع الحد إلى 1,500 RPD.\n"
                    f"تفاصيل الخطأ: {e}"
                ) from e

            if attempt >= max_retries:
                break

            backoff = base_backoff_seconds * (3 ** (attempt - 1))
            if any(x in err_low for x in ('429', 'quota', 'rate', 'resource_exhausted',
                                          'too many', 'unavailable', '503', '500')):
                backoff *= 2
            backoff = min(backoff, 180.0)

            time.sleep(backoff)

    raise RuntimeError(
        f"فشل الاتصال بـ Gemini بعد {max_retries} محاولات للدفعة "
        f"{batch_index + 1}/{total_batches} من ماركة «{brand_name}». "
        f"البحث الإلزامي محفوظ — لم يتم التنازل عنه. آخر خطأ: {last_err}"
    )


# ═════════════════════════════════════════════════════════════════════════════
#  build_output_excel — توافق صارم مع منصة سلة
# ═════════════════════════════════════════════════════════════════════════════

SALLA_MANDATORY = {
    'نوع المنتج':         'منتج جاهز',
    'النوع':              'منتج',
    'هل يتطلب شحن؟':      'نعم',
    'يتطلب شحن؟':         'نعم',
    'يتطلب شحن':          'نعم',
    'الكمية':             10,
    'الكمية المتوفرة':    10,
    'أقصى كمية لكل عميل': 2,
    'الوزن':              1,
    'وحدة الوزن':         'kg',
    'إخفاء خيار التوصيل': 'لا',
    'حالة المنتج':        'نشط',
}

_NULL_LIKE = {'nan', 'NaN', 'NAN', 'None', 'NONE', 'none',
              '<NA>', '<na>', 'NaT', 'nat', 'null', 'NULL', 'undefined'}


def _clean_cell(v):
    """يُحوّل أي قيمة شبيهة بـ NaN إلى سلسلة فارغة ''."""
    if v is None:
        return ''
    try:
        if pd.isna(v):
            return ''
    except (TypeError, ValueError):
        pass
    sv = str(v).strip()
    if sv in _NULL_LIKE or sv == '':
        return ''
    return v if not isinstance(v, str) else sv


def _norm_hdr(s) -> str:
    """تطبيع رؤوس الأعمدة للمقارنة."""
    s = str(s).strip()
    s = (s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')
           .replace('ى', 'ي').replace('ة', 'ه'))
    s = re.sub(r'\s+', ' ', s)
    return s


def build_output_excel(result: dict, original_df: pd.DataFrame, template_bytes: bytes) -> bytes:
    """يبني ملف Excel متوافقاً 100% مع قالب سلة.

    يحتوي **فقط** على:
      • testers_to_add  → التساتر الجديدة المُقترحة (من Gemini أو شبكة الأمان)
      • missing_products → المنتجات الناقصة (موجودة عند المنافسين، غير موجودة لديك)

    لا يحتوي على:
      • products_updated → المنتجات الحالية في متجرك (لا تُكرّر إضافتها)
    """
    brand_col = get_brand_col(original_df)
    name_col  = find_col(original_df, 'name')
    price_col = find_col(original_df, 'price')
    desc_col  = find_col(original_df, 'description')
    cat_col   = find_col(original_df, 'category')
    img_col   = find_col(original_df, 'images')

    brand_name = result.get('brand', '')
    all_cols = list(original_df.columns)

    # 🛡️ مرشّح دفاعي: هياكل أسماء كل المنتجات الموجودة فعلاً في القائمة الأصلية
    # (سواء كانت عطور أساسية أو تساتر). نُسقط أي صف من المخرجات يطابق هياكل
    # موجودة، حتى لو فات ذلك على Gemini أو filter_duplicates.
    existing_skeletons: set = set()
    if name_col and name_col in original_df.columns:
        for raw in original_df[name_col].dropna().astype(str):
            sk = _normalize_perfume_name(_strip_tester_keyword(raw))
            if sk:
                existing_skeletons.add(sk)

    def get_safe_row(base_id):
        if not base_id or 'No.' not in original_df.columns:
            return None
        match = original_df[original_df['No.'].astype(str) == str(base_id)]
        return match.iloc[0] if not match.empty else None

    def apply_salla_mandatory(nr: dict) -> dict:
        for c in all_cols:
            cs = str(c).strip()
            ns = _norm_hdr(cs)
            matched = False
            for key, val in SALLA_MANDATORY.items():
                if cs == key or ns == _norm_hdr(key):
                    nr[c] = val
                    matched = True
                    break
            if matched:
                continue
            if 'نوع المنتج' in cs:
                nr[c] = 'منتج جاهز'
            elif cs == 'النوع':
                nr[c] = 'منتج'
            elif 'يتطلب شحن' in cs:
                nr[c] = 'نعم'
            # ⚠️ مهم: «إخفاء خيار تحديد الكمية» يحتوي كلمة «الكمية» — لذا يجب
            # فحصه قبل قاعدة الكمية حتى لا يأخذ القيمة 10 ويرفضه سلة.
            elif 'إخفاء خيار' in cs or 'اخفاء خيار' in ns:
                nr[c] = 'لا'
            elif 'اقصي كميه' in ns or 'اقصى كميه' in ns:
                nr[c] = 2
            elif ('الكمية' in cs or 'الكميه' in ns) and 'اقصي' not in ns and 'اقصى' not in ns:
                nr[c] = 10
            elif cs == 'الوزن':
                nr[c] = 1
            elif 'وحدة الوزن' in cs or 'وحده الوزن' in ns:
                nr[c] = 'kg'
            elif 'الماركة' in cs and brand_col and c == brand_col:
                nr[c] = brand_name
        return nr

    def _clean_category(cat: str) -> str:
        if not cat:
            return cat
        parts = [p.strip() for p in str(cat).split(',') if p.strip()]
        if not parts:
            return cat
        hierarchical = [p for p in parts if '>' in p]
        return hierarchical[0] if hierarchical else max(parts, key=len)

    rows = []
    skipped_existing = 0  # عدّاد للصفوف التي أُسقطت لأنها موجودة فعلاً
    skipped_sets = 0      # عدّاد للأطقم المتجاهَلة

    for t in result.get('testers_to_add', []):
        t_name = str(t.get('name', '') or '')

        # 🛑 لا تُضِف تستراً لطقم/مجموعة (إن مرّ من Gemini بالخطأ)
        if is_set(t_name):
            skipped_sets += 1
            continue

        # 🛡️ حماية نهائية: تخطّ التساتر التي تطابق منتجاً موجوداً (Gemini ربما
        # اقترحت تستراً لمنتج له تستر مماثل في القائمة الأصلية).
        t_sk = _normalize_perfume_name(_strip_tester_keyword(t_name))
        if t_sk and t_sk in existing_skeletons:
            # نتخطى فقط إذا كان التستر *الأصلي* موجوداً — لا نريد أن نفقد
            # التساتر الجديدة لعطور موجودة بدون تساتر. لذا نتحقق أن المنتج
            # الموجود فعلاً تستر، وليس عطراً أساسياً.
            existing_is_tester = False
            if name_col and name_col in original_df.columns:
                for raw in original_df[name_col].dropna().astype(str):
                    raw_sk = _normalize_perfume_name(_strip_tester_keyword(raw))
                    if raw_sk == t_sk and is_tester(raw):
                        existing_is_tester = True
                        break
            if existing_is_tester:
                skipped_existing += 1
                continue

        nr = {c: '' for c in all_cols}
        base_r = get_safe_row(t.get('base_product_id'))

        if name_col:  nr[name_col]  = _clean_cell(t.get('name', ''))
        if price_col: nr[price_col] = _clean_cell(t.get('new_price', 0)) or 0
        # 📝 نظّف الـ HTML من الفراغات الزائدة لمنع أسطر فارغة في عرض سلة
        if desc_col:
            raw_desc = _clean_cell(t.get('new_description', ''))
            # إذا كان الوصف مازال يحوي placeholders (لم يُملأ بعد)، املأ ما يمكن برمجياً
            desc_str = str(raw_desc) if raw_desc else ''
            if desc_str and ('[اسم الماركة]' in desc_str or '[اسم العطر]' in desc_str):
                # الفولباك: استبدل placeholders البسيطة على الأقل
                bp_cat_for_fill = ''
                if base_r is not None and cat_col:
                    bp_cat_for_fill = str(_clean_cell(base_r.get(cat_col, '')))
                base_name = t.get('_base_perfume_name') or t.get('name', '')
                desc_str = fill_tester_template_basics(
                    desc_str,
                    brand_name=brand_name,
                    perfume_name=base_name,
                    size_ml=int(t.get('size_ml') or 100),
                    base_category=bp_cat_for_fill,
                )
            nr[desc_col] = minify_html(desc_str) if desc_str else ''
        if brand_col: nr[brand_col] = brand_name

        # 🏷️ تصنيف التستر: يُحدَّد دائماً من تصنيفات التستر المخصصة (وليس
        # تصنيف العطر الأساسي). نستخدم تصنيف العطر الأساسي فقط كمدخل
        # لتحديد ما إذا كان رجالي/نسائي/نيش.
        if cat_col:
            base_cat_val = ''
            if base_r is not None:
                base_cat_val = str(_clean_cell(base_r.get(cat_col, '')))
            if not base_cat_val:
                # خذ من _base_category إن وُجد (من شبكة الأمان)
                base_cat_val = str(t.get('_base_category', '') or '')
            t_name = t.get('name', '') or t.get('_base_perfume_name', '')
            nr[cat_col] = map_to_tester_category(base_cat_val, t_name)

        if img_col:
            # 🖼️ احتفظ بقائمة الصور الكاملة مفصولة بفواصل (Salla يدعم متعددة)
            img = _clean_cell(t.get('image_url', ''))
            if not img and base_r is not None:
                raw_img = _clean_cell(base_r.get(img_col, ''))
                if raw_img:
                    # نظّف القائمة من الفراغات والقيم الفارغة
                    img_list = [u.strip() for u in str(raw_img).split(',') if u.strip()]
                    img = ','.join(img_list)
            nr[img_col] = img

        nr = apply_salla_mandatory(nr)
        rows.append(pd.Series(nr))

    for m in result.get('missing_products', []):
        m_name = str(m.get('name', '') or '')

        # 🛑 لا تُضِف طقماً/مجموعة من missing_products
        if is_set(m_name):
            skipped_sets += 1
            continue

        # 🛡️ حماية نهائية: تخطّ المنتجات الناقصة المطابقة لمنتج موجود فعلاً.
        m_sk = _normalize_perfume_name(_strip_tester_keyword(m_name))
        if m_sk and m_sk in existing_skeletons:
            skipped_existing += 1
            continue

        nr = {c: '' for c in all_cols}

        if name_col:  nr[name_col]  = _clean_cell(m.get('name', ''))
        if price_col: nr[price_col] = _clean_cell(m.get('price', 0)) or 0
        if desc_col:
            raw_desc = _clean_cell(m.get('description', ''))
            nr[desc_col] = minify_html(str(raw_desc)) if raw_desc else ''
        if brand_col: nr[brand_col] = _clean_cell(m.get('brand', '')) or brand_name

        cat_val = _clean_cell(m.get('category', ''))
        if cat_col:
            if cat_val:
                nr[cat_col] = _clean_category(cat_val)
            elif not original_df[cat_col].dropna().empty:
                nr[cat_col] = _clean_category(original_df[cat_col].dropna().mode().iloc[0])
            else:
                nr[cat_col] = 'العطور'

        if img_col:
            img1 = _clean_cell(m.get('image_url_1', ''))
            img2 = _clean_cell(m.get('image_url_2', ''))
            imgs = [u for u in (img1, img2) if u]
            nr[img_col] = ','.join(str(u) for u in imgs)

        nr = apply_salla_mandatory(nr)
        rows.append(pd.Series(nr))

    output_df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=all_cols)
    output_df = output_df.fillna('')
    for null_token in _NULL_LIKE:
        output_df = output_df.replace(null_token, '')
    # pandas ≥ 2.1 أزال DataFrame.applymap لصالح DataFrame.map
    if hasattr(output_df, 'map') and callable(getattr(pd.DataFrame, 'map', None)):
        try:
            output_df = output_df.map(_clean_cell)
        except (TypeError, AttributeError):
            output_df = output_df.applymap(_clean_cell)
    else:
        output_df = output_df.applymap(_clean_cell)

    wb = load_workbook(io.BytesIO(template_bytes))

    active_title = wb.active.title
    for sheet_name in list(wb.sheetnames):
        if sheet_name != active_title:
            wb.remove(wb[sheet_name])

    ws = wb.active

    header_row = 2
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        if any(cell and ('اسم' in str(cell) or cell == 'No.') for cell in row):
            header_row = i
            break
    data_start = header_row + 1

    template_headers = [
        ws.cell(row=header_row, column=c).value
        for c in range(1, ws.max_column + 1)
    ]

    semantic_for_template = {
        'تصنيف المنتج': 'category', 'فئة المنتج': 'category', 'فئة': 'category',
        'التصنيف': 'category',
        'اسم المنتج': 'name', 'الاسم': 'name',
        'سعر المنتج': 'price', 'السعر': 'price', 'السعر العادي': 'price',
        'الوصف': 'description', 'وصف المنتج': 'description',
        'صورة المنتج': 'images', 'الصور': 'images', 'صور المنتج': 'images',
        'الماركة': 'brand', 'العلامة التجارية': 'brand',
    }
    semantic_to_dfcol = {
        'category': cat_col, 'name': name_col, 'price': price_col,
        'description': desc_col, 'images': img_col, 'brand': brand_col,
    }

    col_map: dict = {}
    for t_idx, t_hdr in enumerate(template_headers):
        if not t_hdr:
            continue
        t_str = str(t_hdr).strip()
        t_norm = _norm_hdr(t_str)
        matched = None
        for df_col in output_df.columns:
            if t_str == str(df_col).strip() or t_norm == _norm_hdr(str(df_col)):
                matched = df_col
                break
        if matched is None:
            sem = semantic_for_template.get(t_str) or semantic_for_template.get(t_norm)
            if sem and sem in semantic_to_dfcol and semantic_to_dfcol[sem] in output_df.columns:
                matched = semantic_to_dfcol[sem]
        if matched is not None:
            col_map[t_idx + 1] = matched

    direct_template_values: dict = {}
    for t_idx, t_hdr in enumerate(template_headers):
        if not t_hdr or (t_idx + 1) in col_map:
            continue
        cs = str(t_hdr).strip()
        ns = _norm_hdr(cs)
        matched_val = None
        for k, v in SALLA_MANDATORY.items():
            if cs == k or ns == _norm_hdr(k):
                matched_val = v
                break
        if matched_val is not None:
            direct_template_values[t_idx + 1] = matched_val
        elif 'نوع المنتج' in cs:
            direct_template_values[t_idx + 1] = 'منتج جاهز'
        elif cs == 'النوع':
            direct_template_values[t_idx + 1] = 'منتج'
        elif 'يتطلب شحن' in cs:
            direct_template_values[t_idx + 1] = 'نعم'
        # ⚠️ مهم: «إخفاء خيار تحديد الكمية» يحتوي كلمة «الكمية» — لذا يجب
        # فحصه قبل قاعدة الكمية حتى لا يأخذ القيمة 10 ويرفضه سلة.
        elif 'إخفاء خيار' in cs or 'اخفاء خيار' in ns:
            direct_template_values[t_idx + 1] = 'لا'
        elif 'اقصي كميه' in ns or 'اقصى كميه' in ns:
            direct_template_values[t_idx + 1] = 2
        elif ('الكمية' in cs or 'الكميه' in ns) and 'اقصي' not in ns and 'اقصى' not in ns:
            direct_template_values[t_idx + 1] = 10
        elif cs == 'الوزن':
            direct_template_values[t_idx + 1] = 1
        elif 'وحدة الوزن' in cs or 'وحده الوزن' in ns:
            direct_template_values[t_idx + 1] = 'kg'
        elif 'الماركة' in cs:
            direct_template_values[t_idx + 1] = brand_name

    last_written = data_start - 1
    for r_idx, (_, row) in enumerate(output_df.iterrows()):
        excel_row = data_start + r_idx
        last_written = excel_row
        for t_col, df_col in col_map.items():
            val = _clean_cell(row.get(df_col, ''))
            ws.cell(row=excel_row, column=t_col, value=val)
        for t_col, val in direct_template_values.items():
            ws.cell(row=excel_row, column=t_col, value=val)

    WIPE_UNTIL = max(ws.max_row + 1, last_written + 200)
    for r in range(last_written + 1, WIPE_UNTIL):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c, value=None)

    for r in range(data_start, last_written + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            sv = str(cell.value).strip()
            if sv in _NULL_LIKE:
                cell.value = ''

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── INIT STATE ──────────────────────────────────────────────────────────────

def init_state():
    defaults = {
        'df': None,
        'brand_col': None,
        'brands_list': [],
        'filtered_brands': [],
        'current_brand_idx': 0,
        'brand_results': {},
        'processing': False,
        'waiting_confirm': False,
        'current_result': None,
        'template_bytes': None,
        'api_key': os.environ.get('GEMINI_API_KEY', ''),
        'model_name': 'gemini-2.5-flash',
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_state()

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ الإعدادات")

    api_key = st.text_input(
        "🔑 Gemini API Key",
        value=st.session_state.api_key,
        type="password",
        placeholder="AIza...",
        help="احصل على مفتاحك من Google AI Studio",
    )
    st.session_state.api_key = api_key

    # ⭐ زر اختبار المفتاح — يكشف 403 وأخطاء أخرى قبل بدء المعالجة الطويلة
    if api_key:
        if st.button("🔬 اختبار المفتاح", use_container_width=True):
            with st.spinner("جاري اختبار المفتاح..."):
                ok, msg = test_gemini_key(
                    api_key,
                    st.session_state.get('model_name', 'gemini-2.5-flash'),
                )
                if ok:
                    st.success(msg)
                else:
                    st.error(msg)

    # 📋 قائمة نماذج Gemini (محدّثة آخر تحديث: مايو 2026 من توثيق Google الرسمي)
    # https://ai.google.dev/gemini-api/docs/models
    #
    # 🌟 الموصى به افتراضياً: gemini-2.5-flash
    #    - مستقر ومضمون (ليس Preview)
    #    - متاح في الـ Free Tier (10 RPM, 250 RPD)
    #    - متوازن: 90% من جودة Pro بسعر/سرعة Flash
    #
    # ⚡ للسرعة والحجم العالي: gemini-2.5-flash-lite (15 RPM, 1000 RPD)
    # 🧠 للدقة العالية والبحث المعقد: gemini-2.5-pro (5 RPM, 100 RPD)
    # 🚀 الأحدث (Preview، يتطلب Tier 1 المدفوع): gemini-3-flash, gemini-3.1-pro-preview
    model_name = st.selectbox(
        "🤖 النموذج",
        [
            # ─── Stable / Production-ready (موصى بها) ───
            'gemini-2.5-flash',           # ⭐ الأفضل توازناً (الافتراضي)
            'gemini-2.5-flash-lite',      # أسرع وأرخص لكن أقل دقة
            'gemini-2.5-pro',             # أعلى دقة، حصة أقل (5 RPM)
            # ─── Latest aliases (تتحدّث تلقائياً) ───
            'gemini-flash-latest',        # alias لأحدث Flash مستقر
            'gemini-pro-latest',          # alias لأحدث Pro مستقر
            # ─── Preview (يتطلب Cloud Billing) ───
            'gemini-3-flash',             # أحدث جيل، سرعة عالية
            'gemini-3.1-pro-preview',     # أدق نموذج (Preview)
            'gemini-3-flash-lite',        # الأرخص في الجيل الجديد
        ],
        index=0,  # 2.5-flash هو الافتراضي (الأكثر استقراراً)
        help=(
            "⭐ 2.5-flash: الموصى به (متوازن، مستقر، Free Tier)\n"
            "⚡ 2.5-flash-lite: أسرع، حصة أكبر (1000 RPD)\n"
            "🧠 2.5-pro: أعلى دقة، حصة أقل (100 RPD)\n"
            "🚀 3-flash / 3.1-pro: أحدث جيل (يتطلب Cloud Billing)"
        ),
    )
    st.session_state.model_name = model_name

    st.divider()
    st.markdown("## 📁 رفع الملفات")

    uploaded_file = st.file_uploader(
        "ملف المنتجات (Excel / CSV)",
        type=['xlsx', 'xls', 'csv'],
        key='products_file',
    )

    template_file = st.file_uploader(
        "قالب سلة الفارغ (اختياري للتصدير)",
        type=['xlsx'],
        key='template_file',
    )

    if template_file:
        st.session_state.template_bytes = template_file.read()
        st.success("✅ تم تحميل القالب")

    if uploaded_file and st.button("📊 تحليل الملف", use_container_width=True, type="primary"):
        with st.spinner("جاري تحليل الملف..."):
            try:
                df = load_products(uploaded_file)
                st.session_state.df = df
                brand_col = get_brand_col(df)
                st.session_state.brand_col = brand_col
                if brand_col:
                    brands = sorted(df[brand_col].dropna().unique().tolist())
                    st.session_state.brands_list = brands
                    st.session_state.filtered_brands = brands
                    st.session_state.current_brand_idx = 0
                    st.session_state.brand_results = {}
                    st.success(f"✅ {len(df):,} منتج | {len(brands):,} ماركة")
                else:
                    st.error("❌ لم يُعثر على عمود الماركة")
            except Exception as e:
                st.error(f"❌ خطأ في قراءة الملف: {e}")

    if st.session_state.brands_list:
        st.divider()
        st.markdown("## 🔍 الفلاتر المتقدمة")

        selected_brands = st.multiselect(
            "الماركات",
            options=st.session_state.brands_list,
            default=[],
            placeholder="كل الماركات",
        )

        df_now = st.session_state.df
        cat_col = find_col(df_now, 'category') if df_now is not None else None

        selected_cats = []
        if cat_col is not None and df_now is not None:
            all_cats = sorted(df_now[cat_col].dropna().astype(str).unique().tolist())
            selected_cats = st.multiselect(
                "الفئات",
                options=all_cats,
                default=[],
                placeholder="كل الفئات",
            )

        price_range = None
        price_col_now = find_col(df_now, 'price') if df_now is not None else None
        if price_col_now is not None and df_now is not None:
            prices = pd.to_numeric(df_now[price_col_now], errors='coerce').dropna()
            if len(prices) > 0:
                min_p, max_p = int(prices.min()), int(prices.max())
                price_range = st.slider(
                    "نطاق السعر (ريال)", min_p, max_p, (min_p, max_p), step=10
                )

        only_with_testers = st.checkbox("ماركات بها تساتر فقط", value=False)
        sort_by = st.selectbox(
            "ترتيب الماركات حسب",
            ['الاسم أبجدياً', 'عدد المنتجات (تنازلي)', 'عدد المنتجات (تصاعدي)']
        )

        if st.button("✅ تطبيق الفلاتر", use_container_width=True):
            df_f = st.session_state.df
            brand_col_f = st.session_state.brand_col
            filtered = st.session_state.brands_list.copy()

            if selected_brands:
                filtered = [b for b in filtered if b in selected_brands]

            if df_f is not None and selected_cats and cat_col:
                cat_df = df_f[df_f[cat_col].astype(str).isin(selected_cats)]
                cat_brands = set(cat_df[brand_col_f].dropna().astype(str).tolist())
                filtered = [b for b in filtered if b in cat_brands]

            if df_f is not None and price_range and price_col_now:
                p_df = df_f[
                    pd.to_numeric(df_f[price_col_now], errors='coerce').between(
                        price_range[0], price_range[1]
                    )
                ]
                p_brands = set(p_df[brand_col_f].dropna().astype(str).tolist())
                filtered = [b for b in filtered if b in p_brands]

            name_col_f = find_col(df_f, 'name')
            if only_with_testers and df_f is not None and name_col_f:
                tester_brands = set()
                for brand in filtered:
                    bd = df_f[df_f[brand_col_f].astype(str).str.contains(
                        re.escape(str(brand)), case=False, na=False
                    )]
                    if bd[name_col_f].apply(is_tester).any():
                        tester_brands.add(brand)
                filtered = [b for b in filtered if b in tester_brands]

            if sort_by == 'عدد المنتجات (تنازلي)' and df_f is not None:
                filtered.sort(
                    key=lambda b: len(df_f[df_f[brand_col_f].astype(str).str.contains(
                        re.escape(str(b)), case=False, na=False
                    )]),
                    reverse=True
                )
            elif sort_by == 'عدد المنتجات (تصاعدي)' and df_f is not None:
                filtered.sort(
                    key=lambda b: len(df_f[df_f[brand_col_f].astype(str).str.contains(
                        re.escape(str(b)), case=False, na=False
                    )])
                )
            else:
                filtered.sort()

            st.session_state.filtered_brands = filtered
            st.session_state.current_brand_idx = 0
            st.session_state.brand_results = {}
            st.session_state.processing = False
            st.session_state.waiting_confirm = False
            st.success(f"✅ {len(filtered)} ماركة بعد التصفية")
            st.rerun()

# ─── MAIN AREA ────────────────────────────────────────────────────────────────
st.markdown("# 🌿 مهووس | معالج المنتجات الذكي")
st.caption("سياسة جديدة: تستر إلزامي لكل عطر بلا تستر · بحث عن سعر المنافس · سدّ الفجوات")

if st.session_state.df is None:
    st.info("👆 ارفع ملف المنتجات من الشريط الجانبي للبدء")
    st.stop()

df = st.session_state.df
filtered_brands = st.session_state.filtered_brands
brand_col = st.session_state.brand_col
current_idx = st.session_state.current_brand_idx
total_brands = len(filtered_brands)

if total_brands == 0:
    st.warning("⚠️ لا توجد ماركات بعد تطبيق الفلاتر — عدّل الفلاتر من الشريط الجانبي")
    st.stop()

completed = len(st.session_state.brand_results)
col_p1, col_p2, col_p3, col_p4 = st.columns([3, 1, 1, 1])
with col_p1:
    st.markdown(f"### التقدم الإجمالي")
    overall_pct = completed / total_brands if total_brands > 0 else 0
    st.progress(overall_pct, text=f"{completed}/{total_brands} ماركة مكتملة ({overall_pct*100:.0f}%)")
with col_p2:
    st.metric("مكتمل", completed)
with col_p3:
    st.metric("متبقي", total_brands - completed)
with col_p4:
    st.metric("إجمالي", total_brands)

with st.expander(f"📋 قائمة الماركات المختارة ({total_brands})", expanded=False):
    name_col = find_col(df, 'name')
    price_col = find_col(df, 'price')
    summary_rows = []
    for i, brand in enumerate(filtered_brands):
        bd = df[df[brand_col].astype(str).str.contains(
            re.escape(str(brand)), case=False, na=False
        )]
        tester_cnt = bd[name_col].apply(is_tester).sum() if name_col else 0
        avg_price = round(
            pd.to_numeric(bd[price_col], errors='coerce').mean(), 0
        ) if price_col else 0
        status_icon = (
            "✅" if brand in st.session_state.brand_results else
            ("🔄" if i == current_idx and st.session_state.processing else
             ("⏸️" if i > current_idx else "⏭️"))
        )
        summary_rows.append({
            '#': i + 1,
            'الماركة': brand,
            'المنتجات': len(bd),
            'التساتر': tester_cnt,
            'متوسط السعر': avg_price,
            'الحالة': status_icon,
        })
    st.dataframe(
        pd.DataFrame(summary_rows),
        use_container_width=True,
        hide_index=True,
        column_config={'متوسط السعر': st.column_config.NumberColumn(format="%.0f ريال")}
    )

st.divider()

if current_idx >= total_brands:
    st.balloons()
    st.success("🎉 تمت معالجة جميع الماركات بنجاح!")
    results_json = json.dumps(st.session_state.brand_results, ensure_ascii=False, indent=2)
    st.download_button(
        "⬇️ تحميل جميع النتائج (JSON)",
        data=results_json.encode('utf-8'),
        file_name=f"mahwous_all_results_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
        mime="application/json",
        use_container_width=True,
        type="primary",
    )
    if st.button("🔄 بدء جلسة جديدة", use_container_width=True):
        for k in ['brand_results', 'current_brand_idx', 'processing',
                  'waiting_confirm', 'current_result']:
            if k in st.session_state:
                del st.session_state[k]
        st.rerun()
    st.stop()

current_brand = filtered_brands[current_idx]
brand_df = df[df[brand_col].astype(str).str.contains(
    re.escape(str(current_brand)), case=False, na=False
)]
name_col  = find_col(df, 'name')
price_col = find_col(df, 'price')
desc_col  = find_col(df, 'description')
tester_count = brand_df[name_col].apply(is_tester).sum() if name_col else 0

avg_price_val = 0
if price_col:
    prices_series = pd.to_numeric(brand_df[price_col], errors='coerce').dropna()
    avg_price_val = int(prices_series.mean()) if len(prices_series) > 0 else 0

st.markdown(f"""
<div class="brand-card">
  <h2 style="margin:0 0 16px 0;">📦 {current_brand}</h2>
  <div style="display:flex; gap:16px; flex-wrap:wrap;">
    <div class="stat-box" style="flex:1; min-width:120px;">
      <div style="font-size:2em; font-weight:bold;">{len(brand_df):,}</div>
      <div style="color:#aaa; font-size:.9em;">إجمالي المنتجات</div>
    </div>
    <div class="stat-box" style="flex:1; min-width:120px;">
      <div style="font-size:2em; font-weight:bold;">{tester_count:,}</div>
      <div style="color:#aaa; font-size:.9em;">التساتر الحالية</div>
    </div>
    <div class="stat-box" style="flex:1; min-width:120px;">
      <div style="font-size:2em; font-weight:bold;">{avg_price_val:,}</div>
      <div style="color:#aaa; font-size:.9em;">متوسط السعر (ريال)</div>
    </div>
    <div class="stat-box" style="flex:1; min-width:120px;">
      <div style="font-size:2em; font-weight:bold;">{current_idx + 1}/{total_brands}</div>
      <div style="color:#aaa; font-size:.9em;">ترتيب الماركة</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if st.session_state.waiting_confirm and st.session_state.current_result:
    result = st.session_state.current_result
    n_testers = len(result.get('testers_to_add', []))
    n_orphans = len(result.get('orphan_testers', []))
    n_missing = len(result.get('missing_products', []))

    # عدّ التساتر التي وُجدت عند منافس مقابل التي أُضيفت تلقائياً
    testers_list = result.get('testers_to_add', []) or []
    n_with_competitor = sum(1 for t in testers_list if t.get('tester_available_in_market'))
    n_auto_added = sum(1 for t in testers_list if t.get('_auto_added'))

    st.success(
        f"✅ اكتملت معالجة **{current_brand}** | "
        f"{n_testers} تستر جديد ({n_with_competitor} موجود عند منافس · "
        f"{n_auto_added} تلقائي بلا منافس) · "
        f"{n_orphans} تستر بلا أساسي · {n_missing} منتج ناقص"
    )

    tabs = st.tabs([
        f"🏷️ التساتر الجديدة ({n_testers})",
        f"⚠️ تساتر بلا عطر أساسي ({n_orphans})",
        f"🔍 المنتجات الناقصة ({n_missing})",
    ])

    with tabs[0]:
        testers = result.get('testers_to_add', [])
        if testers:
            df_t = pd.DataFrame(testers)
            show_cols = [c for c in [
                'name', 'size_ml', 'original_price', 'new_price',
                'competitor_price', 'tester_available_in_market', 'source_store'
            ] if c in df_t.columns]
            st.dataframe(
                df_t[show_cols] if show_cols else df_t,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'name': 'اسم التستر',
                    'size_ml': st.column_config.NumberColumn('الحجم (مل)', format="%d"),
                    'original_price': st.column_config.NumberColumn('السعر الأصلي', format="%.0f ريال"),
                    'new_price': st.column_config.NumberColumn('سعرنا للتستر', format="%.0f ريال"),
                    'competitor_price': st.column_config.NumberColumn('سعر المنافس (مرجعي)', format="%.0f ريال"),
                    'tester_available_in_market': st.column_config.CheckboxColumn('متوفر عند منافس؟'),
                    'source_store': 'المتجر المرجعي',
                }
            )
            st.caption(
                f"ℹ️ {n_with_competitor} من التساتر وُجدت عند منافس سعودي (سعرها مرجعي فقط) · "
                f"{n_auto_added} أُضيفت تلقائياً بسياسة المتجر بدون منافس مرجعي."
            )
            with st.expander("👁️ معاينة أول وصف HTML"):
                if testers[0].get('new_description'):
                    st.markdown(testers[0]['new_description'], unsafe_allow_html=True)
        else:
            st.info("✅ كل منتجاتك لديها تساتر — لا يوجد ناقص")

    with tabs[1]:
        orphans = result.get('orphan_testers', [])
        if orphans:
            st.warning("هذه التساتر موجودة في متجرك لكن ليس لها عطر أساسي:")
            st.dataframe(pd.DataFrame(orphans), use_container_width=True, hide_index=True)
        else:
            st.info("✅ كل التساتر لديها عطر أساسي مطابق")

    with tabs[2]:
        missing = result.get('missing_products', [])
        if missing:
            df_m = pd.DataFrame(missing)
            show_cols = [c for c in ['name', 'category', 'price', 'is_bestseller', 'source_store']
                         if c in df_m.columns]
            st.dataframe(df_m[show_cols] if show_cols else df_m,
                         use_container_width=True, hide_index=True)
        else:
            st.info("✅ لا توجد منتجات ناقصة")

    st.markdown("### تحميل النتائج")
    dl_col1, dl_col2, dl_col3 = st.columns(3)

    with dl_col1:
        safe_brand = re.sub(r'[^\w]', '_', current_brand)
        st.download_button(
            f"JSON — {current_brand}",
            data=json.dumps(result, ensure_ascii=False, indent=2).encode('utf-8'),
            file_name=f"mahwous_{safe_brand}_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
            mime="application/json",
            use_container_width=True,
        )

    with dl_col2:
        if st.session_state.template_bytes:
            try:
                excel_out = build_output_excel(result, df, st.session_state.template_bytes)
                safe_b = re.sub(r'[^\w]', '_', current_brand)
                st.download_button(
                    f"Excel سلة — {current_brand}",
                    data=excel_out,
                    file_name=f"salla_{safe_b}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.caption(f"⚠️ Excel: {e}")
        else:
            st.caption("🔺 ارفع قالب سلة للحصول على Excel")

    with dl_col3:
        if st.button("⏭️ تخطي هذه الماركة", use_container_width=True, type="secondary"):
            st.session_state.current_brand_idx += 1
            st.session_state.waiting_confirm = False
            st.session_state.current_result = None
            st.session_state.processing = False
            st.rerun()

    st.markdown("---")
    if st.button(
        f"✅ تأكيد واستخراج — ثم الانتقال لـ {filtered_brands[current_idx + 1] if current_idx + 1 < total_brands else 'النهاية'}",
        type="primary",
        use_container_width=True,
    ):
        st.session_state.brand_results[current_brand] = result
        try:
            _safe = re.sub(r'[^\w]', '_', current_brand)
            _p = os.path.join(".mahwous_autosave", f"{_safe}.json")
            if os.path.exists(_p):
                os.remove(_p)
        except Exception:
            pass
        st.session_state.current_brand_idx += 1
        st.session_state.waiting_confirm = False
        st.session_state.current_result = None
        st.session_state.processing = False
        st.rerun()

    st.stop()

if not st.session_state.processing:
    if not st.session_state.api_key:
        st.warning("⚠️ أدخل Gemini API Key في الشريط الجانبي أولاً")
        st.stop()

    with st.expander(f"👁️ معاينة منتجات {current_brand}", expanded=True):
        preview_rows = []
        if name_col:
            for _, row in brand_df.head(15).iterrows():
                preview_rows.append({
                    'اسم المنتج': str(row.get(name_col, '')),
                    'السعر': row.get(price_col, '') if price_col else '',
                    'تستر؟': '✅' if is_tester(str(row.get(name_col, ''))) else '',
                })
        if preview_rows:
            st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
        if len(brand_df) > 15:
            st.caption(f"... و {len(brand_df) - 15} منتج آخر")

    col_start, col_skip = st.columns([4, 1])
    with col_start:
        if st.button(
            f"🚀 بدء معالجة {current_brand} ({len(brand_df)} منتج) بالذكاء الاصطناعي",
            type="primary",
            use_container_width=True,
        ):
            st.session_state.processing = True
            st.rerun()
    with col_skip:
        if st.button("⏭️ تخطي", use_container_width=True, type="secondary"):
            st.session_state.current_brand_idx += 1
            st.rerun()

    st.stop()

# ─── ACTIVE PROCESSING ───────────────────────────────────────────────────────
st.markdown(f"""
<div class="section-header">
🔄 جاري معالجة ماركة: {current_brand} — {len(brand_df)} منتج
</div>
""", unsafe_allow_html=True)

brand_lbl  = st.empty()
brand_bar  = st.progress(0)
prod_lbl   = st.empty()
prod_bar   = st.progress(0)
status_msg = st.empty()

writing_dna = extract_writing_dna(df)

brand_lbl.markdown("**الخطوة 1/3:** جاري تحليل أسلوب الكتابة وتجهيز بيانات المنتجات...")
brand_bar.progress(5)

img_col_now = find_col(df, 'images')
cat_col_now = find_col(df, 'category')
brand_col_now = get_brand_col(df)
products_payload = []
if name_col:
    for _, row in brand_df.iterrows():
        raw_img = str(row.get(img_col_now, '') or '') if img_col_now else ''
        # احتفظ بكل الصور كقائمة مفصولة بفواصل (Salla يدعم متعددة)
        img_list = [u.strip() for u in raw_img.split(',') if u.strip()]
        full_img = ','.join(img_list)

        cat_val = str(row.get(cat_col_now, '') or '') if cat_col_now else ''
        brand_val = str(row.get(brand_col_now, '') or '') if brand_col_now else ''

        products_payload.append({
            'id': str(row.get('No.', row.name)),
            'name': str(row.get(name_col, '')),
            'price': float(pd.to_numeric(row.get(price_col, 0), errors='coerce') or 0),
            'description': '',
            'image_url': full_img,  # ← قائمة كاملة الآن (وليس الصورة الأولى فقط)
            'category': cat_val,
            'brand': brand_val,
            'is_tester': is_tester(str(row.get(name_col, ''))),
        })

n = len(products_payload)

BATCH_SIZE = 15
MAX_PARALLEL = 3
batches = [products_payload[i:i + BATCH_SIZE] for i in range(0, n, BATCH_SIZE)] or [[]]
total_batches = len(batches)

brand_bar.progress(10)
brand_lbl.markdown(
    f"**الخطوة 1/3:** {n} منتج → {total_batches} دفعة "
    f"(حجم {BATCH_SIZE} · {MAX_PARALLEL} متوازية)"
)

SAVE_DIR = ".mahwous_autosave"
os.makedirs(SAVE_DIR, exist_ok=True)
safe_brand_key = re.sub(r'[^\w]', '_', current_brand)
autosave_path = os.path.join(SAVE_DIR, f"{safe_brand_key}.json")

accumulated = {}
if os.path.exists(autosave_path):
    try:
        with open(autosave_path, 'r', encoding='utf-8') as f:
            accumulated = json.load(f)
    except Exception:
        accumulated = {}

completed_set = set(accumulated.get('_completed_batch_ids', []))

_api_key_val = st.session_state.api_key
_model_name_val = st.session_state.model_name

status_lock = threading.Lock()
batch_status = {
    i: {
        'state': 'done' if i in completed_set else 'pending',
        'chars': 0,
        'started_at': None,
        'finished_at': None,
        'error': '',
        'mode': 'grounding',
    }
    for i in range(total_batches)
}


def _run_one(b_idx):
    """يُشغّل دفعة واحدة. البحث (Grounding) محفوظ دائماً —
    call_gemini_brand تُعيد المحاولة داخلياً عند الأخطاء العابرة."""
    with status_lock:
        batch_status[b_idx]['state'] = 'running'
        batch_status[b_idx]['started_at'] = time.time()
        batch_status[b_idx]['mode'] = 'grounding'

    def cb(n_chars):
        with status_lock:
            batch_status[b_idx]['chars'] = n_chars

    return call_gemini_brand(
        brand_name=current_brand,
        products=batches[b_idx],
        full_brand_products=products_payload,
        api_key=_api_key_val,
        writing_dna=writing_dna,
        model_name=_model_name_val,
        use_grounding=True,
        batch_index=b_idx,
        total_batches=total_batches,
        progress_cb=cb,
    )


status_panel = st.empty()


def _render_status():
    with status_lock:
        snap = {i: dict(s) for i, s in batch_status.items()}
    rows = []
    now = time.time()
    for i, s in snap.items():
        if s['state'] == 'pending':
            icon, info = '⏸️', 'في الانتظار'
        elif s['state'] == 'running':
            el = int(now - s['started_at']) if s['started_at'] else 0
            mode_lbl = '🌐 بحث' if s['mode'] == 'grounding' else '⚡ بدون بحث'
            info = f"{mode_lbl} · {el}ث · {s['chars']:,} حرف مستلم"
            icon = '🔄'
        elif s['state'] == 'done':
            dur = int((s['finished_at'] or now) - (s['started_at'] or now))
            icon, info = '✅', f'مكتمل في {dur}ث'
        else:
            icon, info = '❌', (s.get('error', '') or '')[:120]
        rows.append({
            'الدفعة': f"{i + 1}/{total_batches}",
            'الحالة': icon,
            'التفاصيل': info,
        })
    with status_panel.container():
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


try:
    pending_ids = [i for i in range(total_batches) if i not in completed_set]
    results_by_idx = {}

    if pending_ids:
        with ThreadPoolExecutor(max_workers=MAX_PARALLEL) as ex:
            future_to_idx = {ex.submit(_run_one, i): i for i in pending_ids}

            while True:
                done_ids = []
                for fut, idx in list(future_to_idx.items()):
                    if fut.done() and idx not in results_by_idx:
                        try:
                            res = fut.result()
                            results_by_idx[idx] = res
                            with status_lock:
                                batch_status[idx]['state'] = 'done'
                                batch_status[idx]['finished_at'] = time.time()
                        except Exception as e:
                            with status_lock:
                                batch_status[idx]['state'] = 'error'
                                batch_status[idx]['error'] = str(e)
                                batch_status[idx]['finished_at'] = time.time()
                            results_by_idx[idx] = e
                        done_ids.append(idx)

                _render_status()
                done_count = sum(
                    1 for i in range(total_batches)
                    if i in completed_set or i in results_by_idx
                )
                brand_bar.progress(min(10 + int(done_count / total_batches * 65), 75))
                prod_bar.progress(done_count / max(total_batches, 1))
                prod_lbl.markdown(f"📦 {done_count}/{total_batches} دفعة مكتملة")

                if len(results_by_idx) == len(future_to_idx):
                    break
                time.sleep(0.7)

    first_err = None
    for idx, res in results_by_idx.items():
        if isinstance(res, Exception):
            first_err = res
            continue
        merged = filter_duplicates(res, products_payload)
        accumulated = merge_batch_results(accumulated, merged)
        completed_set.add(idx)

    accumulated['_completed_batch_ids'] = sorted(completed_set)
    with open(autosave_path, 'w', encoding='utf-8') as f:
        json.dump(accumulated, f, ensure_ascii=False, indent=2)

    if first_err is not None and len(completed_set) < total_batches:
        raise first_err

    # ⭐ شبكة الأمان النهائية — تضمن وجود تستر لكل عطر أساسي بلا تستر
    # حتى لو فاتت Gemini أحدها، نُضيفه تلقائياً بسعر داخلي.
    accumulated = ensure_all_testers_added(accumulated, products_payload)

    # احفظ مرة أخرى بعد التطبيق التلقائي
    with open(autosave_path, 'w', encoding='utf-8') as f:
        json.dump(accumulated, f, ensure_ascii=False, indent=2)

    # 🌟 إثراء التساتر التلقائية — استدعاء Gemini لتعبئة النوتات والتفاصيل
    # المتخصصة (افتتاحية، قلب، قاعدة، عائلة عطرية، أسئلة شائعة) لكل تستر.
    auto_count = sum(
        1 for t in (accumulated.get('testers_to_add') or [])
        if isinstance(t, dict) and t.get('_auto_added')
    )
    if auto_count > 0:
        brand_lbl.markdown(
            f"**الخطوة 3/3:** إثراء أوصاف {auto_count} تستر تلقائي بمعلومات حقيقية..."
        )
        brand_bar.progress(80)

        def _enrich_progress(done, total):
            try:
                pct = 80 + int(15 * done / max(total, 1))
                brand_bar.progress(min(pct, 95))
                prod_lbl.markdown(f"📝 {done}/{total} وصف مُعبّأ بالمعلومات الحقيقية")
            except Exception:
                pass

        try:
            accumulated = enrich_auto_added_testers(
                accumulated,
                brand_name=brand,
                api_key=api_key,
                model_name=model_name,
                batch_size=8,
                progress_cb=_enrich_progress,
            )
            with open(autosave_path, 'w', encoding='utf-8') as f:
                json.dump(accumulated, f, ensure_ascii=False, indent=2)
        except Exception as enrich_err:
            # فشل الإثراء — نستمر بالقالب المُملوء برمجياً
            st.warning(
                f"⚠️ تعذّر إثراء التساتر التلقائية بمعلومات إضافية "
                f"({enrich_err}). سيتم استخدام الوصف الأساسي فقط."
            )

    result = {k: v for k, v in accumulated.items() if not k.startswith('_')}
    # احتفظ بمعلومة الإضافات التلقائية للعرض
    if accumulated.get('_auto_added_testers'):
        result['_auto_added_testers'] = accumulated['_auto_added_testers']

    brand_bar.progress(75)
    prod_bar.progress(0.7)
    brand_lbl.markdown("**الخطوة 3/3:** معالجة النتائج وتطبيق سياسة التستر الإلزامي...")

    n_tst = len(result.get('testers_to_add', []))
    n_orph = len(result.get('orphan_testers', []))
    n_mis = len(result.get('missing_products', []))
    n_auto = len(result.get('_auto_added_testers', []) or [])

    brand_bar.progress(100)
    prod_bar.progress(1.0)
    brand_lbl.markdown(f"✅ **اكتملت معالجة {current_brand}!**")
    prod_lbl.markdown(
        f"✅ {n_tst} تستر جديد (منهم {n_auto} مُضافين تلقائياً) · "
        f"{n_orph} يتيم · {n_mis} ناقص"
    )
    status_msg.success(
        f"🎯 اكتملت المعالجة — {n_tst} تستر جديد "
        f"({n_auto} تلقائي بسياسة المتجر) | "
        f"{n_orph} تستر بلا أساسي | {n_mis} منتج ناقص"
    )

    st.session_state.current_result = result
    st.session_state.waiting_confirm = True
    st.session_state.processing = False
    time.sleep(0.5)
    st.rerun()

except Exception as e:
    err = str(e)
    brand_bar.progress(0)

    err_low = err.lower()

    # 🔴 خطأ 403 — مشروع محجوب (الأكثر شيوعاً مؤخراً)
    if ('403' in err_low and 'denied access' in err_low) or 'permission_denied' in err_low:
        status_msg.error(
            "❌ **المشروع محجوب من Google** (403 PERMISSION_DENIED)\n\n"
            "هذا الخطأ على مستوى حساب Google وليس في التطبيق."
        )
        with st.expander("🛠️ كيف أحلّ هذه المشكلة؟", expanded=True):
            st.markdown("""
**الحل (مرتّب حسب الأولوية):**

1. اذهب إلى [Google AI Studio - API Keys](https://aistudio.google.com/apikey)
2. احذف المفتاح القديم (أيقونة سلة المهملات 🗑️)
3. اضغط **Create API Key** → اختر **Create in new project**
4. انسخ المفتاح الجديد والصقه في الشريط الجانبي
5. اضغط 🔬 **اختبار المفتاح** للتأكد قبل المتابعة

---

**إن استمرّ الخطأ بعد إنشاء مفتاح جديد:**

- فعّل Cloud Billing من [Google Cloud Console - Billing](https://console.cloud.google.com/billing)
- بعض المناطق (منها أجزاء من السعودية) تتطلب Billing مفعّل حتى لو كنت ضمن الحصة المجانية

**لماذا حصل هذا؟**
- المفتاح ربما تسرّب على GitHub وحُجب تلقائياً
- المشروع قد يكون محجوباً على Google Cloud
- منطقتك تتطلب تفعيل Billing
""")
        st.session_state.processing = False

    # ❌ مفتاح مُبلَّغ عنه كمُسرَّب
    elif 'leaked' in err_low or 'reported as leaked' in err_low:
        status_msg.error(
            "❌ **المفتاح مُبلَّغ عنه كمُسرَّب**\n\n"
            "Google كشف المفتاح على GitHub أو موقع عام وحجبه. "
            "أنشئ مفتاحاً جديداً من https://aistudio.google.com/apikey "
            "وتجنّب رفعه لمستودعات عامة."
        )
        st.session_state.processing = False

    # ❌ مفتاح غير صحيح
    elif ('401' in err_low or 'unauthenticated' in err_low
          or 'api_key_invalid' in err_low or 'invalid api' in err_low
          or 'api key' in err_low):
        status_msg.error(
            "❌ **Gemini API Key غير صحيح أو منتهٍ**\n\n"
            "تحقق من المفتاح في الشريط الجانبي، أو أنشئ مفتاحاً جديداً من "
            "https://aistudio.google.com/apikey"
        )
        st.session_state.processing = False

    # ⚠️ تجاوز الحصة
    elif 'quota' in err_low or 'resource_exhausted' in err_low or '429' in err:
        status_msg.error(
            "⚠️ **تجاوزت حصة الاستخدام**\n\n"
            "**الحلول:**\n"
            "• جرّب الموديل `gemini-2.5-flash-lite` (حصة أكبر بـ 4 مرات)\n"
            "• انتظر حتى 10:00 صباحاً بتوقيت السعودية لإعادة الحصة اليومية\n"
            "• فعّل Cloud Billing لرفع الحد إلى 1,500 طلب/يوم"
        )
        st.session_state.processing = False

    # ⚠️ فلتر الأمان حجب الاستجابة
    elif 'safety' in err_low or 'حُجبت' in err or 'blocked' in err_low:
        status_msg.error(
            "⚠️ **فلتر الأمان حجب استجابة Gemini**\n\n"
            "بعض المنتجات قد تحتوي كلمات يعتبرها Gemini حساسة. "
            "جرّب تخطي هذه الماركة أو إعادة المحاولة."
        )
        st.session_state.processing = False

    # ❌ خطأ عام
    else:
        status_msg.error(f"❌ خطأ غير متوقع:\n\n```\n{err}\n```")
        st.session_state.processing = False

    col_retry, col_skip2 = st.columns(2)
    with col_retry:
        if st.button("🔄 إعادة المحاولة", use_container_width=True, type="primary"):
            st.session_state.processing = True
            st.rerun()
    with col_skip2:
        if st.button("⏭️ تخطي هذه الماركة", use_container_width=True):
            st.session_state.current_brand_idx += 1
            st.session_state.processing = False
            st.rerun()
